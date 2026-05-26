"""
Niamito Business Intelligence Dashboard
app.py  ·  Streamlit Community Cloud deployment
GitHub repo: nikatasler-cloud/niamito-dashboard
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import numpy as np
from datetime import timedelta
import warnings
import os
import pathlib
warnings.filterwarnings("ignore")

# ──────────────────────────────────────────────────────────────────────────────
# BRAND
# ──────────────────────────────────────────────────────────────────────────────
BEIGE  = "#EDE3D8"
BROWN  = "#2C1A0E"
LAVEN  = "#B3B8D9"
GREEN  = "#A8D99A"
CORAL  = "#F07A72"
YELLOW = "#EDD96A"
CREAM  = "#F9F4EF"
MID    = "#6b4c30"

def hex_alpha(hex_color, alpha):
    """Convert a 6-digit hex color + float alpha to an rgba() string."""
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"

MARKET_COLORS = {"SI": BROWN, "HR": LAVEN, "DE": GREEN}
SKU_COLORS    = {
    "NIA-OG-250": BROWN,
    "NIA-VN-250": LAVEN,
    "NIA-CH-250": CORAL,
    "NIA-MP-500": YELLOW,
}

# ──────────────────────────────────────────────────────────────────────────────
# PRODUCT CATEGORIES
# ──────────────────────────────────────────────────────────────────────────────
ALL_CATEGORIES = ["Niamito Oatmeal", "Niamito Fresh Meal", "Niamito Meal in a Bottle"]

_CAT_KEYWORDS = {
    # Oatmeal MUST be checked first — its products also contain "hpp" and flavour words
    # that appear in Fresh Meal keywords, so order matters.
    "Niamito Oatmeal":          ["oatmeal", "oat"],
    "Niamito Meal in a Bottle": ["uht", "470", "cocoa", "cookie", "vanilla", "meal in a bottle"],
    "Niamito Fresh Meal":       ["fresh meal", "fresh", "hpp", "390", "apple", "blueberry",
                                  "spinach", "strawberry", "jaffa", "borovn", "jagod",
                                  "\u0161pina\u010d", "jabolč"],
}

CATEGORY_COLORS = {
    "Niamito Fresh Meal":         GREEN,
    "Niamito Oatmeal":            BROWN,
    "Niamito Meal in a Bottle":   LAVEN,
}

def get_product_category(sku_name: str) -> str:
    """Map a product name to one of three Niamito product categories."""
    name = str(sku_name).lower()
    for cat, keywords in _CAT_KEYWORDS.items():
        if any(k in name for k in keywords):
            return cat
    return "Niamito Fresh Meal"  # safe fallback; Oatmeal is now matched explicitly above

# ──────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Niamito · Business Intelligence",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────────────────────────────────────
# CSS
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

/* ─────────────────────────────────────────────────
   BASE
───────────────────────────────────────────────── */
html, body, [class*="css"], .stApp, .stMarkdown, p, span, div, label, button, input {
    font-family: -apple-system, BlinkMacSystemFont, "Inter", "Helvetica Neue", Arial, sans-serif !important;
    -webkit-font-smoothing: antialiased !important;
    -moz-osx-font-smoothing: grayscale !important;
}

/* ── main content text: always dark ─────────── */
.main .stMarkdown p,
.main .stMarkdown span,
.main .stMarkdown div,
.main p,
.main span,
[data-testid="stMain"] p,
[data-testid="stMain"] span,
[data-testid="stMain"] li,
[data-testid="stMain"] small,
[data-testid="stMain"] caption {
    color: #2C1A0E !important;
}
[data-testid="stMain"] h1,
[data-testid="stMain"] h2,
[data-testid="stMain"] h3,
[data-testid="stMain"] h4 {
    color: #1C1008 !important;
}
/* info / warning / caption boxes: force dark text */
[data-testid="stMain"] [data-testid="stAlert"] *,
[data-testid="stMain"] [data-testid="stCaptionContainer"] * {
    color: #2C1A0E !important;
}

/* ─────────────────────────────────────────────────
   APP BACKGROUND
───────────────────────────────────────────────── */
.stApp { background-color: #EEE6DC; }

/* ─────────────────────────────────────────────────
   SIDEBAR
───────────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: #1C1008 !important;
    border-right: 1px solid rgba(255,255,255,0.06) !important;
}
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div,
section[data-testid="stSidebar"] label {
    color: rgba(240,232,220,0.70) !important;
}
section[data-testid="stSidebar"] .stMarkdown small {
    color: rgba(240,232,220,0.40) !important;
    font-size: 10px !important;
}

/* hide sidebar collapse button */
[data-testid="stSidebarCollapseButton"],
[data-testid="stSidebarHeader"] { display: none !important; }

/* ── section labels ─────────────────────── */
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p {
    font-size: 9px !important;
    letter-spacing: 2px !important;
    text-transform: uppercase !important;
    color: rgba(240,232,220,0.35) !important;
    font-weight: 600 !important;
    margin-bottom: 6px !important;
}

/* ── upload drop zone ─────────────────────────────── */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
    background: rgba(255,255,255,0.05) !important;
    border: 1.5px dashed rgba(240,232,220,0.18) !important;
    border-radius: 16px !important;
    transition: all 0.2s ease !important;
    padding: 20px 16px !important;
    text-align: center !important;
}
/* stretch the span wrapper so the button fills full width */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] > span {
    width: 100% !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"]:hover {
    border-color: rgba(240,232,220,0.40) !important;
    background: rgba(255,255,255,0.08) !important;
}
/* the browse button */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {
    background: #FFFFFF !important;
    color: #111008 !important;
    border: none !important;
    border-radius: 10px !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    letter-spacing: 0.1px !important;
    padding: 9px 0 !important;
    width: 100% !important;
    transition: all 0.15s ease !important;
    display: block !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button:hover {
    background: #ffffff !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 12px rgba(0,0,0,0.15) !important;
}
/* hide the icon wrapper inside the button (hides icon + removes its flex gap) */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button span:has(> [data-testid="stIconMaterial"]) {
    display: none !important;
}
/* force button text dark — overrides the broad sidebar rgba text rule */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button p,
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button span,
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button div {
    color: #111008 !important;
    font-weight: 600 !important;
    font-size: 12px !important;
}
/* hide file size / format hint */
[data-testid="stFileUploaderDropzoneInstructions"],
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] small {
    display: none !important;
}

/* ── uploaded file card — force dark text on light background ── */
section[data-testid="stSidebar"] [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"],
section[data-testid="stSidebar"] [data-testid="stFileUploader"] [data-testid="stFileUploaderFileData"],
section[data-testid="stSidebar"] [data-testid="stFileUploader"] [class*="uploadedFile"] {
    background: #FFFFFF !important;
    border-radius: 10px !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] p,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] span,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] small,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] div,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] [class*="uploadedFile"] * {
    color: #1C1008 !important;
    opacity: 1 !important;
}
/* progress bar under uploaded file */
section[data-testid="stSidebar"] [data-testid="stFileUploader"] [role="progressbar"] {
    display: none !important;
}

/* ── date range picker ───────────────────── */
section[data-testid="stSidebar"] [data-testid="stDateInput"] label {
    display: none !important;
}
section[data-testid="stSidebar"] [data-testid="stDateInput"] [data-baseweb="input"] {
    background: rgba(255,255,255,0.05) !important;
    border: 1px solid rgba(240,232,220,0.15) !important;
    border-radius: 12px !important;
    padding: 2px 4px !important;
}
section[data-testid="stSidebar"] [data-testid="stDateInput"] input {
    color: rgba(240,232,220,0.85) !important;
    font-size: 12px !important;
    font-weight: 500 !important;
    text-align: center !important;
    background: transparent !important;
}
section[data-testid="stSidebar"] [data-testid="stDateInput"] input::placeholder {
    color: rgba(240,232,220,0.35) !important;
}

/* ── multiselect ──────────────────────────── */
section[data-testid="stSidebar"] span[data-baseweb="tag"] {
    background: rgba(240,232,220,0.12) !important;
    border: 1px solid rgba(240,232,220,0.20) !important;
    border-radius: 999px !important;
    font-size: 11px !important;
    font-weight: 500 !important;
    padding: 3px 10px !important;
}
section[data-testid="stSidebar"] [data-baseweb="select"] > div {
    background: rgba(255,255,255,0.05) !important;
    border: 1px solid rgba(240,232,220,0.12) !important;
    border-radius: 12px !important;
}

/* ── period radio → pill buttons ─────────── */
section[data-testid="stSidebar"] [data-testid="stRadio"] > div {
    gap: 3px !important;
    display: flex !important;
    flex-direction: column !important;
}
section[data-testid="stSidebar"] label[data-baseweb="radio"] {
    background: rgba(255,255,255,0.04) !important;
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 10px !important;
    padding: 8px 14px !important;
    transition: all 0.15s ease !important;
    cursor: pointer !important;
    width: 100% !important;
}
section[data-testid="stSidebar"] label[data-baseweb="radio"]:has(input:checked) {
    background: rgba(237,221,210,0.13) !important;
    border-color: rgba(240,232,220,0.22) !important;
}
section[data-testid="stSidebar"] label[data-baseweb="radio"] > div:first-child {
    display: none !important;
}

/* ─────────────────────────────────────────────────
   MAIN CONTENT
───────────────────────────────────────────────── */
.block-container {
    padding-top: 1.8rem !important;
    padding-bottom: 3rem !important;
    max-width: 1400px !important;
}

/* ── metric cards ─────────────────────────── */
div[data-testid="metric-container"] {
    background: #F9F4EF;
    border: 1px solid rgba(44,26,14,0.10);
    border-radius: 16px;
    padding: 16px 20px;
    box-shadow: 0 0 0 0.5px rgba(0,0,0,0.06), 0 2px 6px rgba(44,26,14,0.04), 0 8px 20px rgba(44,26,14,0.07);
}
div[data-testid="metric-container"] label {
    color: #6b4c30 !important;
    font-size: 11px !important;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    font-weight: 600 !important;
}
div[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #1C1008 !important;
    font-size: 28px !important;
    font-weight: 700 !important;
    letter-spacing: -0.5px !important;
}
div[data-testid="metric-container"] [data-testid="stMetricDelta"] {
    font-size: 12px !important;
}

/* ── tabs ─────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    background: rgba(44,26,14,0.07) !important;
    border-radius: 12px !important;
    padding: 4px !important;
    border-bottom: none !important;
}
.stTabs [data-baseweb="tab"] {
    color: rgba(44,26,14,0.55);
    border-radius: 8px !important;
    font-size: 13px;
    font-weight: 500;
    padding: 7px 16px;
    background: transparent !important;
    border: none !important;
    transition: all 0.15s ease !important;
}
.stTabs [aria-selected="true"] {
    color: #1C1008 !important;
    font-weight: 600 !important;
    background: #FFFFFF !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.10), 0 1px 1px rgba(0,0,0,0.06) !important;
}

/* ── iOS segmented control — main content radios ─ */
.stMain [data-testid="stRadio"] > div[role="radiogroup"] {
    display: inline-flex !important;
    flex-direction: row !important;
    background: rgba(44,26,14,0.07) !important;
    border-radius: 10px !important;
    padding: 3px !important;
    gap: 2px !important;
}
.stMain label[data-baseweb="radio"] {
    border-radius: 8px !important;
    padding: 5px 14px !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    color: rgba(44,26,14,0.55) !important;
    transition: all 0.15s ease !important;
    cursor: pointer !important;
    border: none !important;
    background: transparent !important;
}
.stMain label[data-baseweb="radio"]:has(input:checked) {
    background: #FFFFFF !important;
    color: #1C1008 !important;
    font-weight: 600 !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.10), 0 1px 1px rgba(0,0,0,0.06) !important;
}
.stMain label[data-baseweb="radio"] > div:first-child { display: none !important; }

/* ── chart captions ───────────────────────── */
p.chart-caption {
    font-size: 12.5px;
    color: #8a6a4a;
    line-height: 1.55;
    margin-top: 6px;
    margin-bottom: 18px;
    max-width: 520px;
}

/* ── demo badge ───────────────────────────── */
.demo-badge {
    background: #EDD96A;
    color: #2C1A0E;
    font-size: 9px;
    font-weight: 700;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    padding: 3px 10px;
    border-radius: 999px;
    vertical-align: middle;
}

/* ── page header ──────────────────────────── */
.page-header {
    font-size: 22px;
    font-weight: 700;
    color: #1C1008;
    letter-spacing: -0.4px;
    margin-bottom: 2px;
}
.context-line {
    font-size: 12px;
    color: #8a6a4a;
    margin-bottom: 20px;
    font-weight: 500;
}

/* ── dataframe toolbar (Download / Search / Fullscreen) ── */
[data-testid="stDataFrame"] [class*="toolbar"],
[data-testid="stDataFrame"] [class*="Toolbar"],
[data-testid="stElementToolbar"],
[data-testid="stElementToolbarButton"] {
    background: #F9F4EF !important;
    border: 1px solid rgba(44,26,14,0.15) !important;
    border-radius: 8px !important;
    box-shadow: 0 2px 8px rgba(44,26,14,0.12) !important;
}
[data-testid="stElementToolbarButton"] button,
[data-testid="stElementToolbarButton"] svg,
[data-testid="stElementToolbar"] button,
[data-testid="stElementToolbar"] svg {
    color: #2C1A0E !important;
    fill: #2C1A0E !important;
    stroke: #2C1A0E !important;
}
[data-testid="stElementToolbarButton"] button:hover {
    background: #EDE3D8 !important;
}
/* tooltip text inside toolbar */
[data-testid="stElementToolbar"] [class*="tooltip"],
[data-testid="stElementToolbar"] span {
    color: #2C1A0E !important;
    background: #F9F4EF !important;
}
</style>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# CAPTION HELPER
# ──────────────────────────────────────────────────────────────────────────────
def caption(text):
    st.markdown(f"<p class='chart-caption'>{text}</p>", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# CHART LAYOUT HELPER
# ──────────────────────────────────────────────────────────────────────────────
def base_layout(title="", height=340, legend_below=True):
    leg = dict(
        orientation="h", y=-0.20, x=0,
        font=dict(size=10, color=BROWN),
        bgcolor="rgba(0,0,0,0)",
    ) if legend_below else dict(font=dict(size=10, color=BROWN), bgcolor="rgba(0,0,0,0)")
    return dict(
        title=dict(text=title, font=dict(color=BROWN, size=13, family="Georgia"), x=0, pad=dict(l=6)),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor=CREAM,
        font=dict(color=BROWN, size=11, family="Arial"),
        height=height,
        margin=dict(l=12, r=12, t=44, b=20),
        legend=leg,
        xaxis=dict(gridcolor="#e8ddd0", gridwidth=1, zeroline=False, linecolor="#c8b89a"),
        yaxis=dict(gridcolor="#e8ddd0", gridwidth=1, zeroline=False, linecolor="#c8b89a"),
        hoverlabel=dict(bgcolor=CREAM, bordercolor=BROWN, font=dict(color=BROWN, size=11)),
    )


# ──────────────────────────────────────────────────────────────────────────────
# DEMO DATA — Niamito realistic structure Jan 2024–Apr 2026
# ──────────────────────────────────────────────────────────────────────────────
@st.cache_data
def generate_demo_data():
    rng = np.random.default_rng(42)

    PRODUCTS = {
        # Fresh Meal (HPP 390g)
        "APL05": {"name": "Fresh Meal Apple 390g",       "price": 4.20, "bpc": 1, "line": "Niamito Fresh Meal"},
        "BLU05": {"name": "Fresh Meal Blueberry 390g",   "price": 4.20, "bpc": 1, "line": "Niamito Fresh Meal"},
        "SPN05": {"name": "Fresh Meal Spinach 390g",     "price": 4.20, "bpc": 1, "line": "Niamito Fresh Meal"},
        "STR05": {"name": "Fresh Meal Strawberry 390g",  "price": 4.20, "bpc": 1, "line": "Niamito Fresh Meal"},
        # Oatmeal
        "BLK05": {"name": "Blueberry Oat",               "price": 3.50, "bpc": 1, "line": "Niamito Oatmeal"},
        "CHK05": {"name": "Choco Oat",                   "price": 3.50, "bpc": 1, "line": "Niamito Oatmeal"},
        "PCK05": {"name": "Peach Oat",                   "price": 3.50, "bpc": 1, "line": "Niamito Oatmeal"},
        # Meal in a Bottle (UHT 470ml)
        "COU05-DE": {"name": "Meal in a Bottle Cocoa 470ml",   "price": 3.80, "bpc": 1, "line": "Niamito Meal in a Bottle"},
        "CKU05-DE": {"name": "Meal in a Bottle Cookie 470ml",  "price": 3.80, "bpc": 1, "line": "Niamito Meal in a Bottle"},
        "VAU05-DE": {"name": "Meal in a Bottle Vanilla 470ml", "price": 3.80, "bpc": 1, "line": "Niamito Meal in a Bottle"},
    }

    weeks = pd.date_range("2024-01-01", "2026-04-28", freq="W-MON")

    # Base weekly bottles per SKU per market
    BASE = {
        "SI": {
            "APL05": 120, "BLU05": 95, "SPN05": 60, "STR05": 80,
            "BLK05": 70, "CHK05": 55, "PCK05": 50,
            "COU05-DE": 0, "CKU05-DE": 0, "VAU05-DE": 0,
        },
        "HR": {
            "APL05": 50, "BLU05": 40, "SPN05": 25, "STR05": 35,
            "BLK05": 30, "CHK05": 25, "PCK05": 20,
            "COU05-DE": 0, "CKU05-DE": 0, "VAU05-DE": 0,
        },
        "DE": {
            "APL05": 0, "BLU05": 0, "SPN05": 0, "STR05": 0,
            "BLK05": 0, "CHK05": 0, "PCK05": 0,
            "COU05-DE": 80, "CKU05-DE": 65, "VAU05-DE": 55,
        },
    }

    primary_rows = []
    for i, wk in enumerate(weeks):
        growth = 1 + i * 0.008
        for mkt, sku_dict in BASE.items():
            for sku, base_bottles in sku_dict.items():
                if base_bottles == 0:
                    continue
                bottles = max(1, int(base_bottles * growth * rng.uniform(0.85, 1.18)))
                price = PRODUCTS[sku]["price"]
                gross = round(bottles * price, 2)
                primary_rows.append({
                    "week": wk,
                    "market": mkt,
                    "sku_id": sku,
                    "sku_name": PRODUCTS[sku]["name"],
                    "cases": 0,
                    "bottles": bottles,
                    "list_price": price,
                    "gross_revenue": gross,
                    "trade_discount": 0.0,
                    "net_revenue": gross,
                    "funnel_type": "3-tier" if mkt in ["SI", "HR"] else "2-tier",
                    "category": PRODUCTS[sku]["line"],
                })

    prim_df = pd.DataFrame(primary_rows)
    prim_df["category"] = prim_df["sku_name"].apply(get_product_category)

    # ── Secondary sales (retail sell-out, SI only, no revenue) ────────────────
    RETAILER_POOL = [f"Retailer_{i:03d}" for i in range(1, 224)]
    so_rows = []
    so_weeks = pd.date_range("2025-01-06", "2026-04-28", freq="W-MON")
    si_skus = [k for k, v in PRODUCTS.items() if v["line"] in ("Niamito Fresh Meal", "Niamito Oatmeal")]
    for wk in so_weeks:
        n_retailers = rng.integers(30, 80)
        chosen = rng.choice(RETAILER_POOL, size=n_retailers, replace=False)
        for ret in chosen:
            sku = rng.choice(si_skus)
            units = int(rng.integers(2, 25))
            so_rows.append({
                "week": wk,
                "market": "SI",
                "retailer": ret,
                "sku_id": sku,
                "sku_name": PRODUCTS[sku]["name"],
                "bottles_sold": units,
                "consumer_price": 0.0,
                "sellout_revenue": 0.0,
                "funnel_type": "3-tier",
                "category": PRODUCTS[sku]["line"],
            })
    so_df = pd.DataFrame(so_rows)
    so_df["category"] = so_df["sku_name"].apply(get_product_category)

    # ── Distributor stock (SI & HR) ───────────────────────
    stock_rows = []
    for mkt in ["SI", "HR"]:
        stock = 0
        for wk in weeks:
            c_in = int(prim_df[(prim_df["market"] == mkt) & (prim_df["week"] == wk)]["bottles"].sum() // 12)
            c_out = int(c_in * rng.uniform(0.72, 0.96))
            stock = max(0, stock + c_in - c_out)
            stock_rows.append({
                "week": wk, "market": mkt,
                "cases_in": c_in, "cases_out": c_out,
                "stock_cases": stock,
                "stock_to_sales": round(stock / max(c_out, 1), 2),
            })
    stock_df = pd.DataFrame(stock_rows)

    # ── Marketing Calendar ────────────────────────────────
    mkt_campaigns = [
        {"id": "MKT-2024-001", "name": "Q1 Digital SI",           "channel": "Digital",     "market": "SI",  "start": "2024-01-15", "end": "2024-02-28", "media_spend": 2500, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2024-002", "name": "Spring Events SI",         "channel": "Events",      "market": "SI",  "start": "2024-03-01", "end": "2024-04-30", "media_spend": 3800, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2024-003", "name": "Mercator Listing Q1",      "channel": "Team",        "market": "SI",  "start": "2024-01-01", "end": "2024-01-01", "media_spend": 0,    "listing_fee": 4500, "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Store",  "window_days": 0},
        {"id": "MKT-2024-004", "name": "Production Campaign",      "channel": "Production",  "market": "SI",  "start": "2024-05-01", "end": "2024-05-31", "media_spend": 1800, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2024-005", "name": "HR Market Launch",         "channel": "Events",      "market": "HR",  "start": "2024-06-01", "end": "2024-06-30", "media_spend": 3200, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2024-006", "name": "DE Launch Digital",        "channel": "Digital-Intl","market": "DE",  "start": "2024-07-01", "end": "2024-08-31", "media_spend": 5500, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2024-007", "name": "@healthysi Collab",        "channel": "Digital",     "market": "SI",  "start": "2024-09-10", "end": "2024-09-24", "media_spend": 700,  "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": "@healthysi", "reach": 32000, "scope": "Market", "window_days": 14},
        {"id": "MKT-2024-008", "name": "Q4 Traditional SI",        "channel": "Traditional", "market": "SI",  "start": "2024-10-01", "end": "2024-12-31", "media_spend": 4200, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2025-001", "name": "Jan Digital Push",         "channel": "Digital",     "market": "SI",  "start": "2025-01-15", "end": "2025-02-15", "media_spend": 2800, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2025-002", "name": "Spring Events HR",         "channel": "Events",      "market": "HR",  "start": "2025-03-01", "end": "2025-04-30", "media_spend": 3500, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2025-003", "name": "@fitnesswelt_de Collab",   "channel": "Digital-Intl","market": "DE",  "start": "2025-04-01", "end": "2025-04-21", "media_spend": 2000, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": "@fitnesswelt_de", "reach": 115000, "scope": "Market", "window_days": 14},
        {"id": "MKT-2025-004", "name": "Production 2025",          "channel": "Production",  "market": "SI",  "start": "2025-05-01", "end": "2025-05-31", "media_spend": 2100, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2025-005", "name": "Konzum Listing HR",        "channel": "Team",        "market": "HR",  "start": "2025-06-01", "end": "2025-06-01", "media_spend": 0,    "listing_fee": 3800, "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Store",  "window_days": 0},
        {"id": "MKT-2025-006", "name": "Q4 Traditional SI",        "channel": "Traditional", "market": "SI",  "start": "2025-10-01", "end": "2025-12-31", "media_spend": 4800, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2026-001", "name": "Q1 Digital 2026",          "channel": "Digital",     "market": "SI",  "start": "2026-01-15", "end": "2026-02-28", "media_spend": 3200, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
        {"id": "MKT-2026-002", "name": "Spring Events 2026",       "channel": "Events",      "market": "SI",  "start": "2026-03-01", "end": "2026-04-30", "media_spend": 4100, "listing_fee": 0,    "trade_disc": 0,   "roas": None, "influencer": None, "reach": None, "scope": "Market", "window_days": 14},
    ]
    mkt_df = pd.DataFrame(mkt_campaigns)
    mkt_df["total_spend"] = mkt_df["media_spend"] + mkt_df["listing_fee"] + mkt_df["trade_disc"]
    mkt_df["attributed_sales"] = 0.0

    return prim_df, so_df, mkt_df, stock_df, PRODUCTS


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL LOADER (for real uploads)
# ──────────────────────────────────────────────────────────────────────────────
def load_excel(file):
    """
    Load Niamito_Master_Tables.xlsx and map all sheets to the app's
    internal dataframe schema.
    Returns (prim_df, so_df, mkt_df, stock_df, PRODUCTS, exp_df).
    """
    def to_float(val):
        """Safely convert a cell value to float, treating '-', '', None as 0."""
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    CHANNEL_TYPE_MAP = {
        # Slovenian originals → English display names
        "Ekipa": "Team",
        "Produkcija": "Production",
        "Digitalni marketing": "Digital",
        "Digitalni marketing AT": "Digital (AT)",
        "Dogodki & promocije": "Events",
        "Tradicionalni marketing": "Traditional",
        "Influencer": "Influencer",
        # English pass-throughs (already correct)
        "Team": "Team",
        "Digital": "Digital",
        "Events": "Events",
        "Traditional": "Traditional",
        "Production": "Production",
        "Digitalni marketing AT": "Digital-Intl",
        "Dogodki & promocije": "Events",
        "Tradicionalni marketing": "Traditional",
    }

    import openpyxl as _opxl, re as _re

    def _cell_val(v):
        """Extract actual value from openpyxl cell — handles IFERROR formula wrappers that
        Google Sheets / Excel use to cache computed values (IMPORTRANGE, DUMMYFUNCTION, etc.).
        Works whether the workbook was opened with data_only=True (returns cached value directly)
        or data_only=False (returns the formula string, from which we extract the fallback value)."""
        if v is None:
            return None
        s = str(v)
        if not s.startswith("="):
            return v  # plain value — return as-is
        # General IFERROR fallback: match the last , "string" or , number before closing paren
        m = _re.search(r',\s*("([^"]*)"|([-]?\d+(?:\.\d+)?))\s*\)\s*$', s)
        if m:
            if m.group(2) is not None:    # quoted string fallback
                return m.group(2)
            try:
                return float(m.group(3))  # numeric fallback
            except Exception:
                pass
        return None  # formula with no usable fallback

    # Use data_only=True so openpyxl returns cached formula results (e.g. =F2*2.98 → 2145.6).
    # The workbook is always a user-supplied Excel file so cached values are present.
    _wb = _opxl.load_workbook(file, data_only=True)

    def _read_ws_rows(ws, header_row: int, data_start: int):
        """Read a worksheet into a list of dicts, extracting values from formula wrappers."""
        headers = [
            (str(_cell_val(cell.value)).replace("\n", " ").strip()
             if _cell_val(cell.value) is not None else f"_col{i}")
            for i, cell in enumerate(ws[header_row])
        ]
        rows = []
        for row in ws.iter_rows(min_row=data_start, values_only=False):
            extracted = [_cell_val(cell.value) for cell in row]
            if any(v is not None for v in extracted):
                rows.append(dict(zip(headers, extracted)))
        return headers, rows

    # Read sheets with row-3 headers (standard sheets)
    raw = {}
    _PLAIN_HDR = ("primary", "secondary")   # these sheets have header in row 1
    for s in _wb.sheetnames:
        try:
            ws = _wb[s]
            _use_plain = any(k in s.lower() for k in _PLAIN_HDR)
            _, rows = _read_ws_rows(ws,
                                    header_row=1 if _use_plain else 3,
                                    data_start=2 if _use_plain else 4)
            if rows:
                df = pd.DataFrame(rows).dropna(how="all")
                raw[s] = df
        except Exception:
            pass

    # Read Expenses sheet with row-1 headers
    exp_raw = None
    for s in _wb.sheetnames:
        if "expense" in s.lower() or "Expense" in s:
            try:
                ws = _wb[s]
                _, rows = _read_ws_rows(ws, header_row=1, data_start=2)
                if rows:
                    exp_raw = pd.DataFrame(rows).dropna(how="all")
            except Exception:
                pass
            break

    def find(keywords):
        for name, df in raw.items():
            if any(k.lower() in name.lower() for k in keywords):
                return df, name
        return None, None

    prod_raw_df, _ = find(["Product"])
    mkt_raw_df, _  = find(["Marketing", "Calendar"])
    prim_raw_df, _ = find(["Primary", "Sales"])

    # ── v9+ format: Primary Sales has headers in row 1, not row 3 ────────────
    # Detect by checking if "Invoice Date" appears in the first cell of row 1.
    for _sn in _wb.sheetnames:
        if "primary" in _sn.lower() and "sales" in _sn.lower():
            _ws_ps = _wb[_sn]
            _r1c1  = _cell_val(_ws_ps.cell(1, 1).value)
            if _r1c1 and "Invoice Date" in str(_r1c1):
                _, _ps_rows = _read_ws_rows(_ws_ps, header_row=1, data_start=2)
                if _ps_rows:
                    prim_raw_df = pd.DataFrame(_ps_rows).dropna(how="all")
            break

    # Find Secondary Sales sheet
    sec_raw_df = None
    sec_name = None
    for name, df in raw.items():
        if "secondary" in name.lower():
            sec_raw_df = df
            sec_name = name
            break

    # Find Sell-out Template sheet
    sot_raw_df = None
    for name, df in raw.items():
        if "sell" in name.lower() and "template" in name.lower():
            sot_raw_df = df
            break
        if "sell" in name.lower() and name != sec_name:
            sot_raw_df = df
            break

    # ── PRODUCTS dict ─────────────────────────────────────────────────────────
    PRODUCTS = {}
    prod_by_ret_name = {}
    if prod_raw_df is not None:
        ret_name_col = next((c for c in prod_raw_df.columns if "Retailer SKU Name" in c), None)
        sku_col      = next((c for c in prod_raw_df.columns if "Internal SKU" in c), None)
        name_col     = next((c for c in prod_raw_df.columns if "Product Name" in c), None)
        for _, r in prod_raw_df.iterrows():
            if pd.isna(r.get(sku_col, None)):
                continue
            sku = str(r[sku_col]).strip()
            prd_name = str(r[name_col]).strip() if name_col else sku
            PRODUCTS[sku] = {"name": prd_name, "price": 0.0, "bpc": 1}
            if ret_name_col and pd.notna(r.get(ret_name_col)):
                key = str(r[ret_name_col]).strip().lower()
                prod_by_ret_name[key] = {"sku_id": sku, "sku_name": prd_name}

    if not PRODUCTS:
        PRODUCTS = {
            "APL05": {"name": "Fresh Meal Apple 390g",       "price": 4.20, "bpc": 1},
            "BLU05": {"name": "Fresh Meal Blueberry 390g",   "price": 4.20, "bpc": 1},
            "SPN05": {"name": "Fresh Meal Spinach 390g",     "price": 4.20, "bpc": 1},
            "STR05": {"name": "Fresh Meal Strawberry 390g",  "price": 4.20, "bpc": 1},
            "BLK05": {"name": "Blueberry Oat",               "price": 3.50, "bpc": 1},
            "CHK05": {"name": "Choco Oat",                   "price": 3.50, "bpc": 1},
            "PCK05": {"name": "Peach Oat",                   "price": 3.50, "bpc": 1},
            "COU05-DE": {"name": "Meal in a Bottle Cocoa 470ml",   "price": 3.80, "bpc": 1},
            "CKU05-DE": {"name": "Meal in a Bottle Cookie 470ml",  "price": 3.80, "bpc": 1},
            "VAU05-DE": {"name": "Meal in a Bottle Vanilla 470ml", "price": 3.80, "bpc": 1},
        }

    # ── Primary Sales → prim_df ───────────────────────────────────────────────
    prim_rows = []
    if prim_raw_df is not None:
        date_col_p = next((c for c in prim_raw_df.columns if "Invoice Date" in c or "Date" in c), None)
        # v9: "Market or Retailer"  |  old: "Market"
        mkt_col_p  = next((c for c in prim_raw_df.columns if "Market" in c), None)
        # v9: "SKU" (full name or code)  |  old: "Internal SKU Code"
        sku_col_p  = next((c for c in prim_raw_df.columns
                           if "Internal SKU Code" in c or "Internal SKU" in c
                           or c.strip() == "SKU"), None)
        name_col_p = next((c for c in prim_raw_df.columns if "Product Name" in c), None)
        # v9: "Quantity (PCS)"  |  old: "Bottles sold"
        btl_col_p  = next((c for c in prim_raw_df.columns
                           if "Bottles sold" in c or "Bottles Sold" in c or "Bottles" in c
                           or "Quantity" in c or "PCS" in c), None)
        gross_col  = next((c for c in prim_raw_df.columns if "Gross Revenue" in c), None)

        # v9 has "Market or Retailer" — retailer name, not market code
        _is_retailer_col = mkt_col_p and "Retailer" in mkt_col_p

        def _infer_market(raw_val: str, sku_str: str) -> str:
            v = raw_val.strip().upper()
            if v in ("SI", "SLO", "SLOVENIJA"):                         return "SI"
            if v in ("HR", "HRV", "HRVATSKA"):                          return "HR"
            if v in ("DE", "DEU", "DEUTSCHLAND"):                       return "DE"
            # Infer from SKU — DE products contain "- DE"
            if "- DE" in sku_str.upper() or sku_str.upper().endswith("DE"): return "DE"
            # Known SI retailers
            if any(k in v for k in ("PREMA","MERCATOR","SPAR","HOFER","PETROL",
                                    "PTIČKA","ŽITO","TUSCH","INTERSPAR SI")):  return "SI"
            # Known HR retailers
            if any(k in v for k in ("KONZUM","STUDENAC","TOMMY","PLODINE",
                                    "KAUFLAND HR","INTERSPAR HR")):             return "HR"
            return "SI"  # default

        def _resolve_sku(raw_sku: str):
            """Return (sku_id, sku_name) from raw value (full name or internal code)."""
            s = raw_sku.strip()
            if s in PRODUCTS:                          # exact internal code match
                return s, PRODUCTS[s]["name"]
            key = s.lower()
            if key in prod_by_ret_name:                # retailer name lookup
                info = prod_by_ret_name[key]
                return info["sku_id"], info["sku_name"]
            for code in PRODUCTS:                      # prefix match (e.g. "APL05 Organic…")
                if s.upper().startswith(code.upper()):
                    return code, PRODUCTS[code]["name"]
            return s, s                                # unknown — use raw as both id and name

        for _, r in prim_raw_df.iterrows():
            _sku_raw = r.get(sku_col_p)
            if _sku_raw is None or (isinstance(_sku_raw, float) and pd.isna(_sku_raw)):
                continue
            sku_str = str(_sku_raw).strip()
            if not sku_str:
                continue

            raw_date = r.get(date_col_p)
            try:
                week = pd.to_datetime(raw_date, errors="coerce")
            except Exception:
                week = pd.NaT

            _mkt_raw = str(r.get(mkt_col_p, "SI")).strip() if mkt_col_p else "SI"
            mkt = _infer_market(_mkt_raw, sku_str)

            sku_id, sku_name = _resolve_sku(sku_str)
            if name_col_p and pd.notna(r.get(name_col_p)):
                sku_name = str(r.get(name_col_p)).strip() or sku_name

            bottles = int(to_float(r.get(btl_col_p)))
            gross   = to_float(r.get(gross_col))

            # If Gross Revenue is blank, estimate from PRODUCTS list price
            if gross == 0 and bottles > 0:
                price = PRODUCTS.get(sku_id, {}).get("price", 0.0)
                gross = round(price * bottles, 2)

            if bottles == 0 and gross == 0:
                continue

            prim_rows.append({
                "week":           week,
                "market":         mkt,
                "sku_id":         sku_id,
                "sku_name":       sku_name,
                "cases":          0,
                "bottles":        bottles,
                "list_price":     round(gross / max(bottles, 1), 4),
                "gross_revenue":  gross,
                "trade_discount": 0.0,
                "net_revenue":    gross,
                "funnel_type":    "3-tier" if mkt in ("SI", "HR") else "2-tier",
            })

    prim_df = pd.DataFrame(prim_rows) if prim_rows else pd.DataFrame(
        columns=["week","market","sku_id","sku_name","cases","bottles",
                 "list_price","gross_revenue","trade_discount","net_revenue","funnel_type"])
    prim_df["category"] = prim_df["sku_name"].apply(get_product_category)

    stock_df = pd.DataFrame(
        columns=["week","market","cases_in","cases_out","stock_cases","stock_to_sales"])

    # ── Secondary Sales → so_df ───────────────────────────────────────────────
    so_rows = []

    def _parse_secondary_row(r, date_col, retail_col, sku_col_s, pcs_col):
        raw_date = r.get(date_col)
        try:
            week = pd.to_datetime(raw_date, errors="coerce")
        except Exception:
            week = pd.NaT
        retailer = str(r.get(retail_col, "")).strip() if retail_col else ""
        sku_raw  = str(r.get(sku_col_s, "")).strip() if sku_col_s else ""
        prod_info = prod_by_ret_name.get(sku_raw.lower(), {"sku_id": sku_raw, "sku_name": sku_raw})
        units = int(to_float(r.get(pcs_col)))
        return {
            "week":            week,
            "market":          "SI",
            "retailer":        retailer,
            "sku_id":          prod_info["sku_id"],
            "sku_name":        prod_info["sku_name"],
            "bottles_sold":    units,
            "consumer_price":  0.0,
            "sellout_revenue": 0.0,
            "funnel_type":     "3-tier",
        }

    if sec_raw_df is not None:
        date_col_s   = next((c for c in sec_raw_df.columns if "DATE" in c.upper() or "Date" in c), None)
        retail_col_s = next((c for c in sec_raw_df.columns if "RETAIL" in c.upper() or "Store" in c or "Retail" in c), None)
        sku_col_s    = next((c for c in sec_raw_df.columns if c.upper() == "SKU" or "SKU" in c), None)
        pcs_col_s    = next((c for c in sec_raw_df.columns if c.upper() == "PCS" or "PCS" in c or "Units" in c), None)
        for _, r in sec_raw_df.iterrows():
            if pd.isna(r.get(pcs_col_s)):
                continue
            row_data = _parse_secondary_row(r, date_col_s, retail_col_s, sku_col_s, pcs_col_s)
            if row_data["bottles_sold"] > 0:
                so_rows.append(row_data)

    # Sell-out Template — merge additional rows
    if sot_raw_df is not None:
        date_col_t  = next((c for c in sot_raw_df.columns if "Date" in c), None)
        store_col_t = next((c for c in sot_raw_df.columns if "Store" in c or "Retail" in c), None)
        mkt_col_t   = next((c for c in sot_raw_df.columns if "Market" in c), None)
        name_col_t  = next((c for c in sot_raw_df.columns if "Product Name" in c), None)
        units_col_t = next((c for c in sot_raw_df.columns if "Units" in c), None)
        rev_col_t   = next((c for c in sot_raw_df.columns if "Revenue" in c), None)
        for _, r in sot_raw_df.iterrows():
            if pd.isna(r.get(units_col_t)):
                continue
            raw_date = r.get(date_col_t)
            try:
                week = pd.to_datetime(raw_date, errors="coerce")
            except Exception:
                week = pd.NaT
            mkt = str(r.get(mkt_col_t, "SI")).strip() if mkt_col_t else "SI"
            if mkt.upper() in ("SLO", "SLOVENIJA"):
                mkt = "SI"
            if mkt not in ("SI", "HR", "DE"):
                mkt = "SI"
            prd_raw = str(r.get(name_col_t, "")).strip() if name_col_t else ""
            prod_info = prod_by_ret_name.get(prd_raw.lower(), {"sku_id": prd_raw, "sku_name": prd_raw})
            units = int(to_float(r.get(units_col_t)))
            rev   = to_float(r.get(rev_col_t)) if rev_col_t else 0.0
            if units > 0:
                so_rows.append({
                    "week":            week,
                    "market":          mkt,
                    "retailer":        str(r.get(store_col_t, "")).strip() if store_col_t else "",
                    "sku_id":          prod_info["sku_id"],
                    "sku_name":        prod_info["sku_name"],
                    "bottles_sold":    units,
                    "consumer_price":  round(rev / units, 2) if units > 0 and rev > 0 else 0.0,
                    "sellout_revenue": round(rev, 2),
                    "funnel_type":     "3-tier" if mkt in ("SI", "HR") else "2-tier",
                })

    so_df = pd.DataFrame(so_rows) if so_rows else pd.DataFrame(
        columns=["week","market","retailer","sku_id","sku_name","bottles_sold",
                 "consumer_price","sellout_revenue","funnel_type"])
    if not so_df.empty:
        so_df["category"] = so_df["sku_name"].apply(get_product_category)
    else:
        so_df["category"] = pd.Series(dtype=str)

    # ── Marketing Calendar → mkt_df ───────────────────────────────────────────
    mkt_rows = []
    # Read directly from openpyxl to avoid duplicate-header overwrite bug
    _ws_mkt = None
    for _sn in _wb.sheetnames:
        if "marketing" in _sn.lower() or "calendar" in _sn.lower():
            _ws_mkt = _wb[_sn]
            break

    if _ws_mkt is not None:
        _mkt_all_rows = list(_ws_mkt.iter_rows(values_only=True))
        _mkt_hdrs = [str(c).strip() if c else f"_col{i}" for i, c in enumerate(_mkt_all_rows[2])]

        # Map header names to indices (first occurrence wins)
        _hmap = {}
        for _i, _h in enumerate(_mkt_hdrs):
            if _h not in _hmap:
                _hmap[_h] = _i

        def _mkt_get(row, key, default=None):
            idx = _hmap.get(key)
            if idx is None:
                return default
            return row[idx] if idx < len(row) else default

        for _row in _mkt_all_rows[3:]:
            if not any(v is not None for v in _row):
                continue
            cid = _mkt_get(_row, "Campaign ID")
            if cid is None:
                continue
            # Spend at column INDEX 6 always (first "TOTAL Spend (€)")
            total_spend_raw = _row[6] if len(_row) > 6 else None
            total = to_float(total_spend_raw)

            raw_type = str(_mkt_get(_row, "Type", "")).strip()
            channel  = CHANNEL_TYPE_MAP.get(raw_type, raw_type)

            mkt_val = str(_mkt_get(_row, "Market", "")).strip()
            if mkt_val.upper() in ("SLO", "SLOVENIJA"):
                mkt_val = "SI"
            if mkt_val not in ("SI", "HR", "DE"):
                mkt_val = "SI"

            start_raw = _mkt_get(_row, "Start Date")
            end_raw   = _mkt_get(_row, "End Date")
            start_str = str(start_raw)[:10] if start_raw else ""
            end_str   = str(end_raw)[:10]   if end_raw   else ""

            inf_val   = _mkt_get(_row, "Influencer Handle")
            reach_val = _mkt_get(_row, "Estimated Reach")
            scope_raw = _mkt_get(_row, "Attribution Scope")
            window_raw= _mkt_get(_row, "Attribution Window (days)")

            if scope_raw and pd.notna(scope_raw):
                scope = "Market" if "market" in str(scope_raw).lower() else "Store"
            else:
                scope = "Market" if channel in ("Digital","Digital-Intl","Events","Traditional") else "Store"

            inf_clean  = str(inf_val).strip() if inf_val and str(inf_val).strip() not in ("None","nan","") else None
            reach_clean = to_float(reach_val) or None

            mkt_rows.append({
                "id":               str(cid).strip(),
                "name":             str(_mkt_get(_row, "Campaign Name", "")).strip(),
                "channel":          channel,
                "market":           mkt_val,
                "start":            start_str,
                "end":              end_str,
                "media_spend":      total,
                "listing_fee":      to_float(_mkt_get(_row, "Listing Fees (€)")),
                "trade_disc":       to_float(_mkt_get(_row, "Trade Discounts (€)")),
                "total_spend":      total,
                "roas":             None,
                "attributed_sales": 0.0,
                "influencer":       inf_clean,
                "reach":            reach_clean,
                "scope":            scope,
                "window_days":      int(to_float(window_raw) or 14),
            })

    mkt_df = pd.DataFrame(mkt_rows) if mkt_rows else pd.DataFrame(
        columns=["id","name","channel","market","start","end","media_spend",
                 "listing_fee","trade_disc","total_spend","roas","attributed_sales",
                 "influencer","reach","scope","window_days"])

    # ── Expenses → exp_df + stock/promo from Excel ───────────────────────────
    exp_rows = []
    # ── Read Cost Allocation section directly from worksheet (own column layout) ─
    # Structure: section header → instruction → col headers (Month|Stock|Promo|Internal) → 12 month rows
    _ALLOC_MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    # Per-month dicts; summed for waterfall totals, kept for future monthly charts
    _alloc = {m: {"stock": 0.0, "promo": 0.0, "internal": 0.0} for m in _ALLOC_MONTHS}
    _xl_stock_val = _xl_promo_val = _xl_internal_val = 0.0

    _ws_exp = None
    for _sn in _wb.sheetnames:
        if "expense" in _sn.lower():
            _ws_exp = _wb[_sn]; break

    if _ws_exp is not None:
        # Find section header row by scanning col A for "COST ALLOCATION"
        _sec_row = None
        for _r in _ws_exp.iter_rows(min_col=1, max_col=1, values_only=False):
            _v = _cell_val(_r[0].value)
            if _v and "COST ALLOCATION" in str(_v).upper():
                _sec_row = _r[0].row; break
        if _sec_row is not None:
            _data_start = _sec_row + 3
            # Detect format by checking if first data cell is a month name (new monthly table)
            # or a type label like "Active Stock" (old single-row format)
            _first_cell = _cell_val(_ws_exp.cell(_data_start, 1).value)
            _is_monthly = str(_first_cell).strip() in _ALLOC_MONTHS if _first_cell else False

            if _is_monthly:
                # ── New format: rows are months, cols 2/3/4 = stock/promo/internal ──
                for _ri in range(_data_start, _data_start + 12):
                    _mo = _cell_val(_ws_exp.cell(_ri, 1).value)
                    if _mo not in _ALLOC_MONTHS:
                        break
                    _alloc[_mo]["stock"]    = to_float(_cell_val(_ws_exp.cell(_ri, 2).value))
                    _alloc[_mo]["promo"]    = to_float(_cell_val(_ws_exp.cell(_ri, 3).value))
                    _alloc[_mo]["internal"] = to_float(_cell_val(_ws_exp.cell(_ri, 4).value))
            else:
                # ── Old format: rows are types (Active Stock / Promo / Internal),
                #    col 3 = Amount (€) possibly stored as "0,00 €" string ──
                def _parse_euro_str(v):
                    """Parse '1.234,56 €' or '1234.56' or float → float."""
                    if v is None: return 0.0
                    s = str(v).replace("€","").replace(" ","").replace(".","").replace(",",".")
                    try: return float(s)
                    except: return 0.0
                for _ri in range(_data_start, _data_start + 10):
                    _type = _cell_val(_ws_exp.cell(_ri, 1).value)
                    if not _type: continue
                    _type_s = str(_type).strip()
                    _amt = _parse_euro_str(_cell_val(_ws_exp.cell(_ri, 3).value))
                    if _type_s == "Active Stock":
                        _xl_stock_val = _amt
                    elif _type_s == "Promo & External":
                        _xl_promo_val = _amt
                    elif _type_s == "Internal Consumption":
                        _xl_internal_val = _amt

            _xl_stock_val    = _xl_stock_val    or sum(v["stock"]    for v in _alloc.values())
            _xl_promo_val    = _xl_promo_val    or sum(v["promo"]    for v in _alloc.values())
            _xl_internal_val = _xl_internal_val or sum(v["internal"] for v in _alloc.values())

    # ── Parse expense data rows (skip allocation section) ────────────────────
    if exp_raw is not None:
        pl_col   = next((c for c in exp_raw.columns if "Product Line" in c), None)
        sku_col_e= next((c for c in exp_raw.columns if c.strip() == "SKU"), None)
        mon_col  = next((c for c in exp_raw.columns if "Month" in c), None)
        prod_col = next((c for c in exp_raw.columns if "Production cost" in c), None)
        log_col  = next((c for c in exp_raw.columns if "Logistics" in c), None)
        mktg_col = next((c for c in exp_raw.columns if "Marketing" in c and "Promo" in c), None)
        if mktg_col is None:
            mktg_col = next((c for c in exp_raw.columns if "Marketing" in c), None)
        _SKIP_PREFIXES = ("📦", "COST ALLOCATION", "Enter ", "Month", "Active", "Promo &", "Internal")
        for _, r in exp_raw.iterrows():
            pl_val = r.get(pl_col)
            if pl_val is None or (isinstance(pl_val, float) and pd.isna(pl_val)):
                continue
            pl_str = str(pl_val).strip()
            # Skip allocation section rows (header, note, column labels, month rows)
            if pl_str.startswith(_SKIP_PREFIXES) or pl_str in _ALLOC_MONTHS:
                continue
            exp_rows.append({
                "product_line":    pl_str,
                "sku":             str(r.get(sku_col_e, "")).strip() if sku_col_e else "",
                "month":           str(r.get(mon_col, "")).strip() if mon_col else "",
                "production_cost": to_float(r.get(prod_col)),
                "logistics":       to_float(r.get(log_col)),
                "marketing_promo": to_float(r.get(mktg_col)),
            })

    exp_df = pd.DataFrame(exp_rows) if exp_rows else pd.DataFrame(
        columns=["product_line","sku","month","production_cost","logistics","marketing_promo"])

    return prim_df, so_df, mkt_df, stock_df, PRODUCTS, exp_df, _xl_stock_val, _xl_promo_val, _xl_internal_val


# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
        <div style='padding: 32px 4px 24px; text-align:center;'>
            <div style='font-family:Georgia,serif; font-size:22px; font-weight:700;
                        color:#EDE3D8; letter-spacing:-0.4px; line-height:1;'>Niamito</div>
            <div style='font-size:9px; color:rgba(237,227,216,0.28); letter-spacing:3px;
                        text-transform:uppercase; margin-top:6px; font-weight:500;'>Business Intelligence</div>
        </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Data source",
        type=["xlsx"],
        help="Upload Niamito_Master_Tables.xlsx from Google Drive",
        key="file_uploader",
    )

    st.markdown("<div style='margin-top:6px;'></div>", unsafe_allow_html=True)

    period = st.sidebar.selectbox(
        "Period",
        ["This year (YTD)", "Last 3 months", "Last 6 months", "All data", "Custom range"],
        index=0,
    )

    custom_range = None
    if period == "Custom range":
        custom_range = st.date_input(
            "Date range",
            value=(pd.Timestamp("2026-01-01").date(), pd.Timestamp("2026-04-28").date()),
            min_value=pd.Timestamp("2024-01-01").date(),
            max_value=pd.Timestamp("2026-12-31").date(),
            format="DD/MM/YYYY",
        )

    st.markdown("<div style='margin-top:4px;'></div>", unsafe_allow_html=True)

    category_filter = st.multiselect(
        "Product Category",
        options=ALL_CATEGORIES,
        default=ALL_CATEGORIES,
    )

    st.markdown(
        "<hr style='border:none; border-top:1px solid rgba(255,255,255,0.06); margin:10px 0 8px;'>",
        unsafe_allow_html=True,
    )

    metric_mode = st.radio(
        "View as",
        options=["Pieces sold", "Revenue (€)"],
        index=0,
        key="metric_mode",
    )


# ──────────────────────────────────────────────────────────────────────────────
# LOAD DATA
# ──────────────────────────────────────────────────────────────────────────────
demo_mode = True
PRODUCTS = {
    "APL05": {"name": "Fresh Meal Apple 390g",       "price": 4.20, "bpc": 1},
    "BLU05": {"name": "Fresh Meal Blueberry 390g",   "price": 4.20, "bpc": 1},
    "SPN05": {"name": "Fresh Meal Spinach 390g",     "price": 4.20, "bpc": 1},
    "STR05": {"name": "Fresh Meal Strawberry 390g",  "price": 4.20, "bpc": 1},
    "BLK05": {"name": "Blueberry Oat",               "price": 3.50, "bpc": 1},
    "CHK05": {"name": "Choco Oat",                   "price": 3.50, "bpc": 1},
    "PCK05": {"name": "Peach Oat",                   "price": 3.50, "bpc": 1},
    "COU05-DE": {"name": "Meal in a Bottle Cocoa 470ml",   "price": 3.80, "bpc": 1},
    "CKU05-DE": {"name": "Meal in a Bottle Cookie 470ml",  "price": 3.80, "bpc": 1},
    "VAU05-DE": {"name": "Meal in a Bottle Vanilla 470ml", "price": 3.80, "bpc": 1},
}
prim_df  = pd.DataFrame(columns=["week","market","sku_id","sku_name","cases","bottles",
                                  "list_price","gross_revenue","trade_discount","net_revenue","funnel_type","category"])
so_df    = pd.DataFrame(columns=["week","market","retailer","sku_id","sku_name","bottles_sold",
                                  "consumer_price","sellout_revenue","funnel_type","category"])
mkt_df   = pd.DataFrame(columns=["id","name","channel","market","start","end","media_spend",
                                  "listing_fee","trade_disc","total_spend","roas","attributed_sales",
                                  "influencer","reach","scope","window_days"])
stock_df = pd.DataFrame(columns=["week","market","cases_in","cases_out","stock_cases","stock_to_sales"])
exp_df   = pd.DataFrame(columns=["product_line","sku","month","production_cost","logistics","marketing_promo"])
xl_stock_val    = 0.0  # Active stock value (€) — from Excel Cost Allocation section
xl_promo_val    = 0.0  # Promo & External cost (€)
xl_internal_val = 0.0  # Internal Consumption cost (€)

# ── Auto-load from same folder first, then fall back to uploader ──────────
_AUTO_PATH = pathlib.Path(__file__).parent / "Niamito_Master_Tables.xlsx"
_data_source = None
if uploaded is not None:
    _data_source = uploaded
elif _AUTO_PATH.exists():
    _data_source = str(_AUTO_PATH)

if _data_source is not None:
    try:
        prim_df, so_df, mkt_df, stock_df, PRODUCTS, exp_df, xl_stock_val, xl_promo_val, xl_internal_val = load_excel(_data_source)
        demo_mode = False
        _src_label = "auto-loaded from folder" if uploaded is None else "uploaded"
        st.sidebar.success(f"✅ Data {_src_label}")
        if not prim_df.empty and "week" in prim_df.columns:
            latest_prim = prim_df["week"].max()
            if pd.notna(latest_prim):
                st.sidebar.caption(
                    f"📦 Primary Sales: up to {latest_prim.strftime('%b %Y')}. "
                    "Replace the Excel file to update."
                )
    except Exception as e:
        st.sidebar.error(f"Could not read file: {e}")

if demo_mode:
    prim_df, so_df, mkt_df, stock_df, PRODUCTS = generate_demo_data()

prim_f  = prim_df.copy()
so_f    = so_df.copy()
mkt_f   = mkt_df.copy()
stock_f = stock_df.copy()

# Snapshot before category filter — used in SKU Performance to show all products
prim_f_nokcat = prim_f.copy()

# ── Apply category filter ─────────────────────────────
if category_filter and len(category_filter) < len(ALL_CATEGORIES):
    if "category" in prim_f.columns:
        prim_f = prim_f[prim_f["category"].isin(category_filter)]
    if "category" in so_f.columns:
        so_f   = so_f[so_f["category"].isin(category_filter)]

# ── Apply period filter ───────────────────────────────
today = pd.Timestamp.today().normalize()
jan1_this_year = pd.Timestamp(f"{today.year}-01-01")

if period == "Custom range":
    if custom_range and len(custom_range) == 2:
        start_ts, end_ts = pd.Timestamp(custom_range[0]), pd.Timestamp(custom_range[1])
        prim_f  = prim_f[(prim_f["week"] >= start_ts) & (prim_f["week"] <= end_ts)]
        so_f    = so_f[(so_f["week"] >= start_ts) & (so_f["week"] <= end_ts)]
        stock_f = stock_f[(stock_f["week"] >= start_ts) & (stock_f["week"] <= end_ts)]
elif period == "This year (YTD)":
    prim_f  = prim_f[prim_f["week"] >= jan1_this_year]
    so_f    = so_f[so_f["week"] >= jan1_this_year]
    stock_f = stock_f[stock_f["week"] >= jan1_this_year]
elif period == "Last 3 months":
    cutoff = today - pd.DateOffset(months=3)
    prim_f  = prim_f[prim_f["week"] >= cutoff]
    so_f    = so_f[so_f["week"] >= cutoff]
    stock_f = stock_f[stock_f["week"] >= cutoff]
elif period == "Last 6 months":
    cutoff = today - pd.DateOffset(months=6)
    prim_f  = prim_f[prim_f["week"] >= cutoff]
    so_f    = so_f[so_f["week"] >= cutoff]
    stock_f = stock_f[stock_f["week"] >= cutoff]
# "All data" → no filter

# ── Apply period filter to marketing calendar ─────────────────────────────
_mkt_start = pd.to_datetime(mkt_f["start"], errors="coerce") if not mkt_f.empty else pd.Series(dtype="datetime64[ns]")
if period == "Custom range" and custom_range and len(custom_range) == 2:
    _ms, _me = pd.Timestamp(custom_range[0]), pd.Timestamp(custom_range[1])
    mkt_f = mkt_f[(_mkt_start >= _ms) & (_mkt_start <= _me)]
elif period == "This year (YTD)":
    mkt_f = mkt_f[_mkt_start >= jan1_this_year]
elif period == "Last 3 months":
    mkt_f = mkt_f[_mkt_start >= (today - pd.DateOffset(months=3))]
elif period == "Last 6 months":
    mkt_f = mkt_f[_mkt_start >= (today - pd.DateOffset(months=6))]
# "All data" → no filter


# ──────────────────────────────────────────────────────────────────────────────
# HEADER
# ──────────────────────────────────────────────────────────────────────────────
col_logo, col_title = st.columns([0.06, 0.94])
with col_title:
    st.markdown("<h1>Niamito · Business Intelligence</h1>", unsafe_allow_html=True)
    st.markdown(
        f"<p style='color:{MID}; font-size:12px; margin-top:-6px;'>"
        f"{'Demo data — upload your Master Tables to see live numbers' if demo_mode else 'Live data'}"
        f"  ·  Period: {period}</p>",
        unsafe_allow_html=True,
    )

st.markdown("<hr>", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# TABS
# ──────────────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "Primary Sales",
    "Secondary Sales",
    "Marketing",
    "Profitability",
    "SKU Performance",
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 · OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
with tab1:

    # ── KPI calculations ─────────────────────────────────────────────────────
    total_gross   = prim_f["gross_revenue"].sum() if not prim_f.empty else 0
    prim_pieces   = prim_f["bottles"].sum() if not prim_f.empty else 0
    avg_price_pc  = total_gross / max(prim_pieces, 1)

    # YoY comparison: same calendar window one year back
    def _prior_year_prim(pf_full, period_name, jan1_ty, today_ts, custom_range_val):
        if period_name == "This year (YTD)":
            py_start = jan1_ty - pd.DateOffset(years=1)
            py_end   = today_ts - pd.DateOffset(years=1)
        elif period_name == "Last 3 months":
            py_start = today_ts - pd.DateOffset(months=3) - pd.DateOffset(years=1)
            py_end   = today_ts - pd.DateOffset(years=1)
        elif period_name == "Last 6 months":
            py_start = today_ts - pd.DateOffset(months=6) - pd.DateOffset(years=1)
            py_end   = today_ts - pd.DateOffset(years=1)
        elif period_name == "Custom range" and custom_range_val and len(custom_range_val) == 2:
            py_start = pd.Timestamp(custom_range_val[0]) - pd.DateOffset(years=1)
            py_end   = pd.Timestamp(custom_range_val[1]) - pd.DateOffset(years=1)
        else:
            return pd.DataFrame()
        if pf_full.empty:
            return pd.DataFrame()
        return pf_full[(pf_full["week"] >= py_start) & (pf_full["week"] <= py_end)]

    py_df     = _prior_year_prim(prim_df, period, jan1_this_year, today,
                                  custom_range if period == "Custom range" else None)
    py_pieces = py_df["bottles"].sum() if not py_df.empty else 0
    py_gross  = py_df["gross_revenue"].sum() if not py_df.empty else 0

    def _delta_str(curr, prev, is_currency=False):
        if prev == 0:
            return None
        diff = curr - prev
        pct  = diff / prev * 100
        sign = "+" if diff >= 0 else ""
        if is_currency:
            return f"{sign}€{abs(diff):,.0f}  ({sign}{pct:.1f}% vs prior yr)"
        return f"{sign}{diff:,.0f}  ({sign}{pct:.1f}% vs prior yr)"

    pieces_delta = _delta_str(prim_pieces, py_pieces)
    gross_delta  = _delta_str(total_gross, py_gross, is_currency=True)

    # ── 2026 run rate ─────────────────────────────────────────────────────────
    _pf_2026_all = prim_f[prim_f["week"].dt.year == 2026] if not prim_f.empty else prim_f
    if not _pf_2026_all.empty:
        _2026_mo_grp = (_pf_2026_all.copy()
                        .assign(_mkey=lambda d: d["week"].dt.to_period("M").astype(str))
                        .groupby("_mkey").agg(pieces=("bottles","sum"), revenue=("gross_revenue","sum"))
                        .reset_index())
        _n_actual_mo = len(_2026_mo_grp)
        _run_rate_pcs = int(_2026_mo_grp["pieces"].mean() * 12)
        _run_rate_rev = _2026_mo_grp["revenue"].mean() * 12
    else:
        _n_actual_mo = 0
        _run_rate_pcs = 0
        _run_rate_rev = 0.0

    # ── KPI row ───────────────────────────────────────────────────────────────
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Pieces Sold", f"{prim_pieces:,.0f}", delta=pieces_delta,
              help=f"Prior year same period: {py_pieces:,.0f} pcs" if py_pieces else None)
    k2.metric("Gross Revenue", f"€{total_gross:,.0f}", delta=gross_delta,
              help=f"Prior year same period: €{py_gross:,.0f}" if py_gross else None)
    k3.metric("Avg Price / Piece", f"€{avg_price_pc:.2f}")
    if _run_rate_rev > 0:
        _rr_label = f"€{_run_rate_rev:,.0f}" if metric_mode != "Pieces sold" else f"{_run_rate_pcs:,}"
        k4.metric("2026 Full-Year Run Rate",
                  _rr_label,
                  help=f"Based on {_n_actual_mo} month(s) of 2026 data × 12. Not a guaranteed forecast.")

    st.markdown("")

    if not demo_mode and not prim_f.empty:
        _latest_inv   = prim_f["week"].max()
        _earliest_inv = prim_f["week"].min()
        st.info(
            f"📦 **Primary sales data:** {_earliest_inv.strftime('%b %Y') if pd.notna(_earliest_inv) else '?'}"
            f" → {_latest_inv.strftime('%b %Y') if pd.notna(_latest_inv) else '?'}"
        )

    # ── Monthly sales by category + 2026 Forecast ────────────────────────────
    col_trend, col_fc = st.columns([0.55, 0.45])

    with col_trend:
        if not prim_f.empty:
            _pf_mo = prim_f.copy()
            _pf_mo["_mkey"] = _pf_mo["week"].dt.to_period("M").astype(str)
            _pf_mo["month"] = _pf_mo["week"].dt.strftime("%b %Y")
            cat_monthly = _pf_mo.groupby(["_mkey", "month", "category"]).agg(
                pieces=("bottles","sum"),
                revenue=("gross_revenue","sum"),
            ).reset_index().sort_values("_mkey")
            y_col   = "pieces" if metric_mode == "Pieces sold" else "revenue"
            y_label = "Pieces" if metric_mode == "Pieces sold" else "Revenue (\u20ac)"
            fig_sw = go.Figure()
            for cat in ALL_CATEGORIES:
                d = cat_monthly[cat_monthly["category"] == cat]
                if d.empty:
                    continue
                fig_sw.add_trace(go.Bar(
                    x=d["month"], y=d[y_col],
                    name=cat,
                    marker_color=CATEGORY_COLORS.get(cat, BROWN),
                    hovertemplate=f"<b>{cat}</b><br>%{{x}}: %{{y:,}}<extra></extra>",
                ))
            l_sw = base_layout(f"Monthly Sales by Category ({y_label})", height=320)
            l_sw["barmode"] = "stack"
            l_sw["xaxis"]["tickangle"] = -30
            if metric_mode != "Pieces sold":
                l_sw["yaxis"]["tickprefix"] = "\u20ac"
            fig_sw.update_layout(**l_sw)
            st.plotly_chart(fig_sw, use_container_width=True)
        else:
            st.info("No primary sales data.")

    with col_fc:
        # 2026 actuals + projected remaining months at YTD run rate
        _fc_y   = "pieces" if metric_mode == "Pieces sold" else "revenue"
        _fc_lbl = "Pieces" if metric_mode == "Pieces sold" else "Revenue (\u20ac)"

        if not _pf_2026_all.empty and _n_actual_mo > 0:
            _fc_act = (_pf_2026_all.copy()
                       .assign(_mkey=lambda d: d["week"].dt.to_period("M").astype(str),
                               month=lambda d: d["week"].dt.strftime("%b %Y"))
                       .groupby(["_mkey","month"]).agg(
                           pieces=("bottles","sum"), revenue=("gross_revenue","sum"))
                       .reset_index().sort_values("_mkey"))

            _actual_mkeys = set(_fc_act["_mkey"])
            _all_2026_mkeys = [f"2026-{m:02d}" for m in range(1, 13)]
            _fc_mkeys = [mk for mk in _all_2026_mkeys if mk not in _actual_mkeys]
            _fc_labels = [pd.Period(mk).strftime("%b %Y") for mk in _fc_mkeys]
            _monthly_avg_fc = _fc_act[_fc_y].mean()

            fig_fc = go.Figure()
            fig_fc.add_trace(go.Bar(
                x=_fc_act["month"], y=_fc_act[_fc_y],
                name="Actual", marker_color=BROWN, opacity=0.9,
                hovertemplate=f"%{{x}}: %{{y:,}} {_fc_lbl} (actual)<extra></extra>",
            ))
            if _fc_labels:
                fig_fc.add_trace(go.Bar(
                    x=_fc_labels, y=[_monthly_avg_fc] * len(_fc_labels),
                    name="Forecast (run rate)", marker_color=LAVEN, opacity=0.55,
                    hovertemplate=f"%{{x}}: %{{y:,}} {_fc_lbl} (projected)<extra></extra>",
                ))
            _fc_total = int(_fc_act[_fc_y].sum() + _monthly_avg_fc * len(_fc_labels))
            l_fc = base_layout(
                f"2026 Monthly Sales + Forecast  ·  Full-year est: "
                f"{'€' if _fc_y == 'revenue' else ''}{_fc_total:,}",
                height=320,
            )
            l_fc["barmode"] = "group"
            l_fc["xaxis"]["tickangle"] = -30
            l_fc["legend"]["orientation"] = "h"
            l_fc["legend"]["y"] = -0.25
            if _fc_y == "revenue":
                l_fc["yaxis"]["tickprefix"] = "\u20ac"
            fig_fc.update_layout(**l_fc)
            st.plotly_chart(fig_fc, use_container_width=True)
        else:
            st.info("No 2026 data yet — forecast will appear once 2026 invoices are loaded.")

    # ── Monthly performance table ─────────────────────────────────────────────
    if not prim_f.empty:
        _mo_tbl = (prim_f.copy()
                   .assign(_mkey=lambda d: d["week"].dt.to_period("M").astype(str),
                           month=lambda d: d["week"].dt.strftime("%b %Y"))
                   .groupby(["_mkey","month"]).agg(
                       pieces=("bottles","sum"), revenue=("gross_revenue","sum"))
                   .reset_index().sort_values("_mkey"))
        _mo_tbl["MoM %"] = _mo_tbl["revenue"].pct_change() * 100
        _mo_tbl = _mo_tbl.sort_values("_mkey", ascending=False)
        _mo_tbl_out = _mo_tbl[["month","pieces","revenue","MoM %"]].copy()
        _mo_tbl_out.columns = ["Month","Pieces","Revenue (€)","MoM %"]
        _mo_tbl_out["Pieces"]      = _mo_tbl_out["Pieces"].apply(lambda x: f"{int(x):,}")
        _mo_tbl_out["Revenue (€)"] = _mo_tbl_out["Revenue (€)"].apply(lambda x: f"€{x:,.0f}")
        _mo_tbl_out["MoM %"]       = _mo_tbl_out["MoM %"].apply(
            lambda x: f"+{x:.1f}%" if pd.notna(x) and x >= 0 else (f"{x:.1f}%" if pd.notna(x) else "—"))
        st.markdown(
            f"<p style='font-size:12px; color:{MID}; margin:4px 0 6px;'>"
            "<b>Monthly Performance</b></p>",
            unsafe_allow_html=True,
        )
        st.dataframe(_mo_tbl_out, use_container_width=True, hide_index=True)

    # ── Sales by category summary table ──────────────────────────────────────
    if not prim_f.empty:
        _cat_summary = prim_f.groupby("category").agg(
            pieces=("bottles", "sum"),
            revenue=("gross_revenue", "sum"),
        ).reset_index().sort_values("revenue", ascending=False)
        _cat_summary["Avg price/pc (€)"] = (_cat_summary["revenue"] / _cat_summary["pieces"].clip(lower=1)).round(2)
        _cat_summary.columns = ["Category", "Pieces", "Revenue (€)", "Avg price/pc (€)"]
        _cat_summary["Pieces"] = _cat_summary["Pieces"].apply(lambda x: f"{int(x):,}")
        _cat_summary["Revenue (€)"] = _cat_summary["Revenue (€)"].apply(lambda x: f"€{x:,.0f}")
        _cat_summary["Avg price/pc (€)"] = _cat_summary["Avg price/pc (€)"].apply(lambda x: f"€{x:.2f}")
        st.markdown(
            f"<p style='font-size:12px; color:{MID}; margin:4px 0 6px;'>"
            "<b>Sales by Product Category</b></p>",
            unsafe_allow_html=True,
        )
        st.dataframe(_cat_summary, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 · SECONDARY SALES
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown("<h2>Secondary Sales — Retailer Sell-Out</h2>", unsafe_allow_html=True)
    st.markdown(
        f"<p style='font-size:12px;color:{MID};margin-top:-6px;'>"
        "Units sold through to end consumers, reported by retail partners.</p>",
        unsafe_allow_html=True,
    )

    if not so_f.empty and "retailer" in so_f.columns:
        _so2 = so_f.copy()
        _so2_units   = int(_so2["bottles_sold"].sum())
        _so2_stores  = _so2["retailer"].nunique()
        _so2_skus    = _so2["sku_name"].nunique()
        _so2_top     = _so2.groupby("retailer")["bottles_sold"].sum().idxmax()
        _so2_top_u   = int(_so2.groupby("retailer")["bottles_sold"].sum().max())

        # YoY comparison for secondary sales
        def _prior_year_so(so_full, period_name, jan1_ty, today_ts, custom_range_val):
            if period_name == "This year (YTD)":
                py_start = jan1_ty - pd.DateOffset(years=1)
                py_end   = today_ts - pd.DateOffset(years=1)
            elif period_name == "Last 3 months":
                py_start = today_ts - pd.DateOffset(months=3) - pd.DateOffset(years=1)
                py_end   = today_ts - pd.DateOffset(years=1)
            elif period_name == "Last 6 months":
                py_start = today_ts - pd.DateOffset(months=6) - pd.DateOffset(years=1)
                py_end   = today_ts - pd.DateOffset(years=1)
            elif period_name == "Custom range" and custom_range_val and len(custom_range_val) == 2:
                py_start = pd.Timestamp(custom_range_val[0]) - pd.DateOffset(years=1)
                py_end   = pd.Timestamp(custom_range_val[1]) - pd.DateOffset(years=1)
            else:
                return pd.DataFrame()
            if so_full.empty:
                return pd.DataFrame()
            return so_full[(so_full["week"] >= py_start) & (so_full["week"] <= py_end)]

        _py_so = _prior_year_so(so_df, period, jan1_this_year, today,
                                 custom_range if period == "Custom range" else None)
        _py_so_units  = int(_py_so["bottles_sold"].sum()) if not _py_so.empty else 0
        _py_so_stores = _py_so["retailer"].nunique() if not _py_so.empty else 0
        _py_so_skus   = _py_so["sku_name"].nunique() if not _py_so.empty else 0

        so2_k1, so2_k2, so2_k3, so2_k4 = st.columns(4)
        so2_k1.metric("Units Sold (sell-out)", f"{_so2_units:,}",
                      delta=_delta_str(_so2_units, _py_so_units) if _py_so_units else None,
                      help=f"Prior year same period: {_py_so_units:,} units" if _py_so_units else None)
        so2_k2.metric("Stores Reporting", f"{_so2_stores}",
                      delta=_delta_str(_so2_stores, _py_so_stores) if _py_so_stores else None)
        so2_k3.metric("SKUs Tracked", f"{_so2_skus}",
                      delta=_delta_str(_so2_skus, _py_so_skus) if _py_so_skus else None)
        so2_k4.metric("Top Store", _so2_top,
                      delta=f"{_so2_top_u:,} units",
                      help="Store with highest sell-out volume")

        st.markdown("<div style='margin-top:14px;'></div>", unsafe_allow_html=True)

        col_so2a, col_so2b = st.columns([0.60, 0.40])

        with col_so2a:
            _store_rank2 = (
                _so2.groupby("retailer")["bottles_sold"].sum()
                .reset_index().sort_values("bottles_sold", ascending=True).tail(20)
            )
            _s2_colors = [BROWN if i % 2 == 0 else LAVEN for i in range(len(_store_rank2))]
            fig_so2 = go.Figure(go.Bar(
                x=_store_rank2["bottles_sold"], y=_store_rank2["retailer"],
                orientation="h",
                marker=dict(color=_s2_colors, line=dict(color=CREAM, width=0.5)),
                text=[f"{int(v):,}" for v in _store_rank2["bottles_sold"]],
                textposition="outside",
                hovertemplate="<b>%{y}</b><br>%{x:,} units<extra></extra>",
            ))
            l_so2 = base_layout("Top 20 Stores by Units Sold", height=max(320, len(_store_rank2) * 24))
            l_so2["xaxis"]["title"] = "Units Sold"
            l_so2["margin"]["l"] = 220
            l_so2["margin"]["r"] = 60
            fig_so2.update_layout(**l_so2)
            st.plotly_chart(fig_so2, use_container_width=True)

        with col_so2b:
            _cat_so2 = (
                _so2.groupby("sku_name")["bottles_sold"].sum()
                .reset_index().sort_values("bottles_sold", ascending=False)
            )
            _cat_so2_colors = [CATEGORY_COLORS.get(get_product_category(n), BROWN) for n in _cat_so2["sku_name"]]
            fig_mix2 = go.Figure(go.Pie(
                labels=_cat_so2["sku_name"], values=_cat_so2["bottles_sold"],
                hole=0.50,
                marker=dict(colors=_cat_so2_colors, line=dict(color=CREAM, width=2)),
                textinfo="percent",
                hovertemplate="<b>%{label}</b><br>%{value:,} units  (%{percent})<extra></extra>",
            ))
            l_mix2 = base_layout("Sell-Out Mix by Product", height=320, legend_below=False)
            l_mix2["legend"] = dict(orientation="v", x=1.02, y=0.5,
                                    font=dict(size=9, color=BROWN), bgcolor="rgba(0,0,0,0)")
            fig_mix2.update_layout(**l_mix2)
            st.plotly_chart(fig_mix2, use_container_width=True)

        # All-stores table
        st.markdown(
            f"<p style='font-size:12px;font-weight:600;color:{BROWN};margin:8px 0 6px;'>"
            "All Stores — Sell-Out Detail</p>",
            unsafe_allow_html=True,
        )
        _store_tbl2 = (
            _so2.groupby(["retailer","market"] if "market" in _so2.columns else ["retailer"])
            .agg(units=("bottles_sold","sum"), skus=("sku_name","nunique"))
            .reset_index().sort_values("units", ascending=False)
            .rename(columns={"retailer":"Store","market":"Market","units":"Units Sold","skus":"SKUs"})
        )
        _store_tbl2["Units Sold"] = _store_tbl2["Units Sold"].apply(lambda x: f"{int(x):,}")
        st.dataframe(_store_tbl2, use_container_width=True, hide_index=True)

    else:
        st.info("No secondary sales data available. Upload a file with Secondary Sales sheet data.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 · MARKETING
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    ch_palette = [BROWN, LAVEN, GREEN, CORAL, YELLOW, MID, "#c8b89a"]

    # ── KPI row ───────────────────────────────────────────────────────────────
    total_spend_mkt = mkt_f["total_spend"].sum() if not mkt_f.empty else 0
    n_campaigns     = len(mkt_f) if not mkt_f.empty else 0
    avg_spend_camp  = total_spend_mkt / max(n_campaigns, 1)

    _mkt_gross_rev = prim_f["gross_revenue"].sum() if not prim_f.empty else 0
    _roi_per_eur   = _mkt_gross_rev / max(total_spend_mkt, 1)

    mk1, mk2, mk3, mk4 = st.columns(4)
    mk1.metric("Total Spend",          f"\u20ac{total_spend_mkt:,.0f}")
    mk2.metric("Campaigns",            f"{n_campaigns}")
    mk3.metric("Avg Spend / Campaign", f"\u20ac{avg_spend_camp:,.0f}")
    mk4.metric("Revenue per € Spent",  f"€{_roi_per_eur:.2f}",
               help="Gross revenue from primary sales ÷ total marketing spend. Shows how much revenue each €1 of marketing generates.")

    st.markdown("<hr style='margin:8px 0 16px'>", unsafe_allow_html=True)

    if mkt_f.empty:
        st.info("No marketing data available.")
    else:
        # ── Key insight callout ───────────────────────────────────────────────
        if not mkt_f.empty and total_spend_mkt > 0:
            _top_ch     = mkt_f.groupby("channel")["total_spend"].sum().idxmax()
            _top_ch_pct = mkt_f.groupby("channel")["total_spend"].sum().max() / total_spend_mkt * 100
            _top_mkt    = mkt_f.groupby("market")["total_spend"].sum().idxmax()
            _top_mkt_pct = mkt_f.groupby("market")["total_spend"].sum().max() / total_spend_mkt * 100
            _n_months   = mkt_f["start"].apply(lambda x: pd.to_datetime(x, errors="coerce")).dt.to_period("M").nunique()
            _avg_mo     = total_spend_mkt / max(_n_months, 1)
            st.markdown(
                f"<div style='background:#f0ebe3;border-left:4px solid {BROWN};padding:10px 16px;"
                f"border-radius:6px;margin-bottom:14px;font-size:13px;color:#3a2e24;line-height:1.9;'>"
                f"📊 <b>Top channel:</b> {_top_ch} — {_top_ch_pct:.0f}% of total spend &nbsp;·&nbsp; "
                f"🌍 <b>Top market:</b> {_top_mkt} — {_top_mkt_pct:.0f}% of budget &nbsp;·&nbsp; "
                f"📅 <b>Avg monthly spend:</b> €{_avg_mo:,.0f} across {_n_months} active months"
                f"</div>",
                unsafe_allow_html=True,
            )

        # ── Secondary sell-out vs Marketing spend ─────────────────────────────
        if not so_f.empty:
            _so_mkt_mo = so_f.copy()
            _so_mkt_mo["_mkey"] = _so_mkt_mo["week"].dt.to_period("M").astype(str)
            _so_mkt_mo["month"] = _so_mkt_mo["week"].dt.strftime("%b %Y")
            _so_units_mo = (_so_mkt_mo.groupby(["_mkey","month"])["bottles_sold"]
                            .sum().reset_index().sort_values("_mkey"))

            _mk_mo2 = mkt_f.copy()
            _mk_mo2["_mkey"] = pd.to_datetime(_mk_mo2["start"], errors="coerce").dt.to_period("M").astype(str)
            _mk_mo2["month"] = pd.to_datetime(_mk_mo2["start"], errors="coerce").dt.strftime("%b %Y")
            _mk_spend_mo2 = (_mk_mo2.groupby(["_mkey","month"])["total_spend"]
                             .sum().reset_index().sort_values("_mkey"))

            _so_mkt_merged = _so_units_mo.merge(
                _mk_spend_mo2[["_mkey","month","total_spend"]], on=["_mkey","month"], how="outer"
            ).sort_values("_mkey").fillna(0)

            fig_so_mkt = go.Figure()
            fig_so_mkt.add_trace(go.Bar(
                x=_so_mkt_merged["month"], y=_so_mkt_merged["bottles_sold"],
                name="Units Sold (secondary)", marker_color=GREEN, opacity=0.85, yaxis="y",
                hovertemplate="%{x}: %{y:,} units<extra></extra>",
            ))
            fig_so_mkt.add_trace(go.Scatter(
                x=_so_mkt_merged["month"], y=_so_mkt_merged["total_spend"],
                name="Marketing Spend", mode="lines+markers",
                line=dict(color=CORAL, width=2),
                marker=dict(size=7, color=CORAL),
                yaxis="y2",
                hovertemplate="%{x}: \u20ac%{y:,.0f} spend<extra></extra>",
            ))
            l_so_mkt = base_layout("Secondary Units Sold vs Marketing Spend", height=300)
            l_so_mkt["yaxis"]["title"] = dict(text="Units Sold", font=dict(size=10))
            l_so_mkt["yaxis2"] = dict(
                overlaying="y", side="right", tickprefix="\u20ac",
                showgrid=False, tickfont=dict(color=CORAL, size=10),
            )
            l_so_mkt["xaxis"]["tickangle"] = -30
            l_so_mkt["legend"]["orientation"] = "h"
            l_so_mkt["legend"]["y"] = -0.25
            fig_so_mkt.update_layout(**l_so_mkt)
            st.plotly_chart(fig_so_mkt, use_container_width=True)

        # ── Channel breakdown (full width) ────────────────────────────────────
        col_ch_bar = st.container()
        with col_ch_bar:
            ch_spend = mkt_f.groupby("channel").agg(
                spend=("total_spend","sum"),
                n=("id","count"),
            ).reset_index()
            ch_spend = ch_spend[ch_spend["spend"] > 0].sort_values("spend", ascending=True)
            fig_ch = go.Figure(go.Bar(
                x=ch_spend["spend"],
                y=ch_spend["channel"],
                orientation="h",
                marker=dict(color=ch_palette[:len(ch_spend)],
                            line=dict(color=CREAM, width=0.8)),
                text=[f"\u20ac{v:,.0f}  ·  {int(n)} campaign{'s' if n!=1 else ''}"
                      for v, n in zip(ch_spend["spend"], ch_spend["n"])],
                textposition="outside",
                hovertemplate="<b>%{y}</b><br>\u20ac%{x:,.0f}<extra></extra>",
            ))
            l_ch = base_layout("Spend & Campaigns by Channel", height=300)
            l_ch["xaxis"]["tickprefix"] = "\u20ac"
            l_ch["margin"]["l"] = 120
            l_ch["margin"]["r"] = 180
            fig_ch.update_layout(**l_ch)
            st.plotly_chart(fig_ch, use_container_width=True)

        # ── Row 2: Monthly spend (stacked) + Cumulative burn curve ────────────
        col_monthly, col_cum = st.columns(2)

        _mk_mo = mkt_f.copy()
        _mk_mo["_mkey"] = pd.to_datetime(_mk_mo["start"], errors="coerce").dt.to_period("M").astype(str)
        _mk_mo["month"] = pd.to_datetime(_mk_mo["start"], errors="coerce").dt.strftime("%b %Y")
        _monthly_ch = (_mk_mo.groupby(["_mkey","month","channel"])["total_spend"]
                       .sum().reset_index().sort_values("_mkey"))
        _monthly_tot = (_mk_mo.groupby(["_mkey","month"])["total_spend"]
                        .sum().reset_index().sort_values("_mkey"))
        _monthly_tot["cumulative"] = _monthly_tot["total_spend"].cumsum()

        with col_monthly:
            fig_mo = go.Figure()
            for idx, ch in enumerate(_monthly_ch["channel"].unique()):
                d = _monthly_ch[_monthly_ch["channel"] == ch]
                fig_mo.add_trace(go.Bar(
                    x=d["month"], y=d["total_spend"],
                    name=ch,
                    marker_color=ch_palette[idx % len(ch_palette)],
                    hovertemplate=f"<b>{ch}</b><br>%{{x}}<br>\u20ac%{{y:,.0f}}<extra></extra>",
                ))
            l_mo = base_layout("Monthly Spend by Channel", height=320)
            l_mo["barmode"] = "stack"
            l_mo["yaxis"]["tickprefix"] = "\u20ac"
            l_mo["xaxis"]["tickangle"] = -35
            l_mo["legend"]["y"] = -0.30
            fig_mo.update_layout(**l_mo)
            st.plotly_chart(fig_mo, use_container_width=True)

        with col_cum:
            fig_cum = go.Figure()
            fig_cum.add_trace(go.Bar(
                x=_monthly_tot["month"], y=_monthly_tot["total_spend"],
                name="Monthly spend",
                marker_color=LAVEN, opacity=0.7,
                hovertemplate="%{x}<br>\u20ac%{y:,.0f}<extra></extra>",
            ))
            fig_cum.add_trace(go.Scatter(
                x=_monthly_tot["month"], y=_monthly_tot["cumulative"],
                name="Cumulative",
                mode="lines+markers",
                line=dict(color=BROWN, width=2.5),
                marker=dict(size=6, color=BROWN),
                yaxis="y2",
                hovertemplate="%{x}<br>Cumulative: \u20ac%{y:,.0f}<extra></extra>",
            ))
            l_cum = base_layout("Budget Burn Rate (cumulative)", height=320)
            l_cum["yaxis"]["tickprefix"] = "\u20ac"
            l_cum["yaxis2"] = dict(
                overlaying="y", side="right", showgrid=False,
                tickprefix="\u20ac",
                tickfont=dict(color=BROWN, size=10),
            )
            l_cum["xaxis"]["tickangle"] = -35
            l_cum["legend"]["y"] = -0.30
            fig_cum.update_layout(**l_cum)
            st.plotly_chart(fig_cum, use_container_width=True)

        # ── Sales vs Marketing Investment ─────────────────────────────────────
        if not prim_f.empty:
            st.markdown("<h2>Sales vs Marketing Investment</h2>", unsafe_allow_html=True)
            _mkt_pmo = mkt_f.copy()
            _mkt_pmo["_mkey"] = pd.to_datetime(_mkt_pmo["start"], errors="coerce").dt.to_period("M").astype(str)
            _mkt_pmo["month"] = pd.to_datetime(_mkt_pmo["start"], errors="coerce").dt.strftime("%b %Y")
            _mkt_spend_mo = _mkt_pmo.groupby(["_mkey","month"])["total_spend"].sum().reset_index().sort_values("_mkey")

            _prim_pmo = prim_f.copy()
            _prim_pmo["_mkey"] = _prim_pmo["week"].dt.to_period("M").astype(str)
            _prim_pmo["month"] = _prim_pmo["week"].dt.strftime("%b %Y")
            _prim_rev_mo = _prim_pmo.groupby(["_mkey","month"]).agg(
                revenue=("gross_revenue","sum"), pieces=("bottles","sum")
            ).reset_index().sort_values("_mkey")

            _mkt_merged = _prim_rev_mo.merge(
                _mkt_spend_mo[["_mkey","month","total_spend"]], on=["_mkey","month"], how="outer"
            ).sort_values("_mkey").fillna(0)
            _mkt_merged["roi"] = (_mkt_merged["revenue"] / _mkt_merged["total_spend"].clip(lower=1)).round(2)

            _mkt_si_y   = "pieces" if metric_mode == "Pieces sold" else "revenue"
            _mkt_si_lbl = "Pieces" if metric_mode == "Pieces sold" else "Revenue (€)"

            fig_mkt_corr = go.Figure()
            fig_mkt_corr.add_trace(go.Bar(
                x=_mkt_merged["month"], y=_mkt_merged[_mkt_si_y],
                name=_mkt_si_lbl, marker_color=LAVEN, opacity=0.8, yaxis="y",
                hovertemplate=f"%{{x}}: %{{y:,}} {_mkt_si_lbl}<extra></extra>",
            ))
            fig_mkt_corr.add_trace(go.Scatter(
                x=_mkt_merged["month"], y=_mkt_merged["total_spend"],
                name="Marketing Spend", mode="lines+markers",
                line=dict(color=CORAL, width=2),
                marker=dict(size=7, color=CORAL),
                yaxis="y2",
                hovertemplate="%{x}: \u20ac%{y:,.0f} spend<extra></extra>",
            ))
            l_mkt_corr = base_layout("Monthly Sales vs Marketing Spend", height=320)
            l_mkt_corr["yaxis2"] = dict(
                overlaying="y", side="right", tickprefix="\u20ac",
                showgrid=False, tickfont=dict(color=CORAL, size=10),
            )
            if metric_mode != "Pieces sold":
                l_mkt_corr["yaxis"]["tickprefix"] = "\u20ac"
            l_mkt_corr["xaxis"]["tickangle"] = -30
            l_mkt_corr["legend"]["orientation"] = "h"
            l_mkt_corr["legend"]["y"] = -0.25
            fig_mkt_corr.update_layout(**l_mkt_corr)
            st.plotly_chart(fig_mkt_corr, use_container_width=True)

        # ── All campaigns table ───────────────────────────────────────────────
        st.markdown(
            f"<h2>All Campaigns</h2>"
            f"<p style='font-size:11px;color:{MID};margin-top:-6px;'>"
            f"{len(mkt_f)} campaigns · \u20ac{mkt_f['total_spend'].sum():,.0f} total in selected period</p>",
            unsafe_allow_html=True,
        )
        _tbl_cols = [c for c in ["id","name","channel","market","start","end","total_spend","roas"]
                     if c in mkt_f.columns]
        _tbl = (mkt_f[_tbl_cols].copy()
                .assign(_start_dt=lambda d: pd.to_datetime(d["start"], errors="coerce"))
                .sort_values("_start_dt", ascending=False)
                .drop(columns=["_start_dt"])
                ).rename(columns={
            "id": "ID", "name": "Campaign", "channel": "Channel",
            "market": "Mkt", "start": "Start", "end": "End",
            "total_spend": "Spend (\u20ac)", "roas": "ROAS",
        })
        if "Spend (\u20ac)" in _tbl.columns:
            _tbl["Spend (\u20ac)"] = _tbl["Spend (\u20ac)"].apply(lambda x: f"\u20ac{x:,.0f}")
        if "ROAS" in _tbl.columns:
            _tbl["ROAS"] = _tbl["ROAS"].apply(lambda x: f"{x:.1f}\u00d7" if pd.notna(x) else "—")
        st.dataframe(_tbl, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 · PROFITABILITY
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown("<h2>Profitability — From Revenue to Gross Margin</h2>", unsafe_allow_html=True)
    st.markdown(
        f"<p style='font-size:12px; color:{MID};'>"
        "Gross revenue minus costs gives the gross margin. "
        "Each bar below shows how much of the top line survives after each deduction.</p>",
        unsafe_allow_html=True,
    )

    # ── Context banner ────────────────────────────────────────────────────────
    if not demo_mode:
        _MO_ORDER = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        _exp_months_raw = exp_df["month"].unique().tolist() if not exp_df.empty else []
        _exp_months_sorted = sorted(_exp_months_raw, key=lambda m: _MO_ORDER.index(m[:3]) if m[:3] in _MO_ORDER else 99)
        _exp_range = (f"{_exp_months_sorted[0]} – {_exp_months_sorted[-1]}"
                      if len(_exp_months_sorted) >= 2
                      else (_exp_months_sorted[0] if _exp_months_sorted else "n/a"))
        _prim_2026_pieces = prim_f[prim_f["week"].dt.year == 2026]["bottles"].sum() if not prim_f.empty else 0
        _raw_prod_cost = exp_df["production_cost"].sum() if not exp_df.empty else 0
        _cost_per_pc = (_raw_prod_cost / max(_prim_2026_pieces, 1)) if not exp_df.empty else 0
        st.markdown(
            f"""<div style='background:#f0ebe3;border-left:4px solid {BROWN};padding:10px 16px;
            border-radius:6px;margin-bottom:12px;font-size:13px;color:#3a2e24;line-height:1.9;'>
            📅 <b>Revenue:</b> 2026 primary invoices only &nbsp;·&nbsp;
            💰 <b>Costs:</b> Expenses {_exp_range} 2026 (all 12 months of annual cost plan) &nbsp;·&nbsp;
            📦 <b>Units sold (2026):</b> {int(_prim_2026_pieces):,} pcs &nbsp;·&nbsp;
            🔩 <b>Full-year prod. cost/pc:</b> €{_cost_per_pc:.2f}<br>
            📐 <b>Formula:</b> Gross Revenue − Production COGS − Logistics − Mktg &amp; Promo − Promo &amp; External − Internal Consumption = Gross Margin
            </div>""",
            unsafe_allow_html=True,
        )

    _avail_mkt = sorted(prim_f["market"].unique().tolist()) if not prim_f.empty else []
    _wf_opts = ["All markets"] + _avail_mkt
    wf_mkt = st.radio("Market:", options=_wf_opts, horizontal=True, key="wf_mkt")

    stock_val    = xl_stock_val
    promo_val    = xl_promo_val
    internal_val = xl_internal_val

    # Profitability uses 2026 revenue only (expenses are 2026-only; mixing years distorts margins)
    _prim_2026 = prim_f[prim_f["week"].dt.year == 2026] if not prim_f.empty else prim_f
    p_data = _prim_2026 if wf_mkt == "All markets" else (
        _prim_2026[_prim_2026["market"] == wf_mkt] if not _prim_2026.empty else _prim_2026)
    m_data = mkt_f if wf_mkt == "All markets" else (
        mkt_f[mkt_f["market"].isin([wf_mkt, "ALL"])] if not mkt_f.empty else mkt_f)

    gross      = p_data["gross_revenue"].sum() if not p_data.empty else 0
    trade_disc = p_data["trade_discount"].sum() if not p_data.empty else 0
    mkt_spend  = m_data["total_spend"].sum() if not m_data.empty else 0

    # Pull expense totals
    prod_cost  = exp_df["production_cost"].sum() if not exp_df.empty else 0
    logistics  = exp_df["logistics"].sum()        if not exp_df.empty else 0
    mktg_promo = exp_df["marketing_promo"].sum()  if not exp_df.empty else 0

    # ── COGS allocation ───────────────────────────────────────────────────────
    # stock_val  → balance sheet asset (capitalized, NOT expensed this period)
    # promo_val  → Promo & External spend (expensed, separate P&L line on top of COGS)
    # internal_val → Internal Consumption (expensed, separate P&L line on top of COGS)
    cogs_adjusted = max(prod_cost - stock_val, 0)
    total_exp = cogs_adjusted + promo_val + internal_val + logistics + mktg_promo

    # Gross Margin = Revenue − COGS − all additional expense lines
    gross_margin = gross - cogs_adjusted - promo_val - internal_val - logistics - mktg_promo

    # ── Inventory / allocation callout ────────────────────────────────────────
    if stock_val > 0:
        st.markdown(
            f"<div style='background:#fdf6ee;border-left:3px solid {BROWN};padding:7px 14px;"
            f"border-radius:5px;margin-bottom:8px;font-size:12.5px;color:#3a2e24;'>"
            f"🏭 <b>Active stock value:</b> €{stock_val:,.2f} — removed from P&L, sits on the balance sheet as inventory."
            f"</div>",
            unsafe_allow_html=True,
        )

    # ── Waterfall ─────────────────────────────────────────────────────────────
    if total_exp > 0:
        wf_labels  = ["Gross Revenue", "Production COGS", "Logistics", "Mktg & Promo"]
        wf_measure = ["absolute",       "relative",        "relative",  "relative"]
        wf_values  = [gross,            -cogs_adjusted,    -logistics,  -mktg_promo]
        # Always show Promo & External and Internal Consumption when expense data is present
        wf_labels.append("Promo & External")
        wf_measure.append("relative")
        wf_values.append(-promo_val)
        wf_labels.append("Internal Consumption")
        wf_measure.append("relative")
        wf_values.append(-internal_val)
        wf_labels.append("Gross Margin")
        wf_measure.append("total")
        wf_values.append(gross_margin)
    else:
        wf_labels  = ["Gross Revenue", "Marketing Spend", "Gross Margin"]
        wf_measure = ["absolute",       "relative",        "total"]
        gross_margin = gross - mkt_spend
        wf_values  = [gross,            -mkt_spend,        gross_margin]

    fig_wf = go.Figure(go.Waterfall(
        name="Value Leakage",
        orientation="v",
        measure=wf_measure,
        x=wf_labels,
        y=wf_values,
        text=[f"\u20ac{abs(v):,.0f}" for v in wf_values],
        textposition="outside",
        connector=dict(line=dict(color=MID, width=1, dash="dot")),
        increasing=dict(marker=dict(color=GREEN)),
        decreasing=dict(marker=dict(color=CORAL)),
        totals=dict(marker=dict(color=CORAL if gross_margin < 0 else GREEN)),
        hovertemplate="<b>%{x}</b><br>\u20ac%{y:,.0f}<extra></extra>",
    ))
    margin_pct = gross_margin / max(gross, 1) * 100
    layout_wf = base_layout(
        f"Profitability Waterfall — {wf_mkt}  ·  Gross Margin: {margin_pct:.1f}%",
        height=400,
    )
    layout_wf["yaxis"]["tickprefix"] = "\u20ac"
    # Ensure the y-axis always shows negative territory when gross_margin < 0
    _wf_y_min = min(gross_margin, 0) * 1.25
    _wf_y_max = gross * 1.15
    if _wf_y_min < 0:
        layout_wf["yaxis"]["range"] = [_wf_y_min, _wf_y_max]
        layout_wf["yaxis"]["zeroline"] = True
        layout_wf["yaxis"]["zerolinecolor"] = BROWN
        layout_wf["yaxis"]["zerolinewidth"] = 2
        # Add a visible zero-baseline annotation
        fig_wf.add_hline(y=0, line_dash="solid", line_color=BROWN, line_width=2, opacity=0.6)
    fig_wf.update_layout(**layout_wf)
    st.plotly_chart(fig_wf, use_container_width=True)

    # ── Summary table ─────────────────────────────────────────────────────────
    wf_tbl = pd.DataFrame({
        "Line": wf_labels,
        "Amount (\u20ac)": [f"\u20ac{abs(v):,.0f}" for v in wf_values],
        "% of Gross":  [f"{abs(v)/max(gross,1)*100:.1f}%" for v in wf_values],
        "Cumulative (\u20ac)": [
            f"\u20ac{(gross + sum(wf_values[1:i+1])):,.0f}" if i > 0 else f"\u20ac{gross:,.0f}"
            for i in range(len(wf_values))
        ],
    })
    st.dataframe(wf_tbl, use_container_width=True, hide_index=True)

    # ── Revenue vs Cost summary ───────────────────────────────────────────────
    if not prim_f.empty or not exp_df.empty:
        _pf_summ_gross = prim_f["gross_revenue"].sum() if not prim_f.empty else 0
        _pf_summ_pcs   = prim_f["bottles"].sum() if not prim_f.empty else 0
        _exp_total     = (exp_df["production_cost"].sum() + exp_df["logistics"].sum()
                          + exp_df["marketing_promo"].sum()) if not exp_df.empty else 0
        _pf4_k1, _pf4_k2, _pf4_k3, _pf4_k4 = st.columns(4)
        _pf4_k1.metric("Gross Revenue",   f"€{_pf_summ_gross:,.0f}")
        _pf4_k2.metric("Total Costs",     f"€{_exp_total:,.0f}")
        _pf4_k3.metric("Gross Margin",    f"€{_pf_summ_gross - _exp_total:,.0f}")
        _pf4_k4.metric("Margin %",        f"{(_pf_summ_gross - _exp_total)/max(_pf_summ_gross,1)*100:.1f}%")
        st.markdown("")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 · SKU PERFORMANCE
# ══════════════════════════════════════════════════════════════════════════════
with tab5:

    st.markdown("<h2>SKU Performance</h2>", unsafe_allow_html=True)

    if not prim_f_nokcat.empty:
        sku_prim = prim_f_nokcat.groupby("sku_id").agg(
            sku_name=("sku_name","first"),
            bottles=("bottles","sum"),
            gross=("gross_revenue","sum"),
            net=("net_revenue","sum"),
        ).reset_index().sort_values("gross", ascending=False)

        # ── Gross Revenue by SKU — horizontal ranked bar ───────────────────────
        _sku_rev_sorted = sku_prim.sort_values("gross", ascending=True)
        _sku_rev_colors = [CATEGORY_COLORS.get(get_product_category(n), BROWN)
                           for n in _sku_rev_sorted["sku_name"]]
        fig_rev = go.Figure(go.Bar(
            x=_sku_rev_sorted["gross"],
            y=_sku_rev_sorted["sku_name"],
            orientation="h",
            marker=dict(color=_sku_rev_colors),
            text=[f"\u20ac{v:,.0f}" for v in _sku_rev_sorted["gross"]],
            textposition="outside",
            hovertemplate="<b>%{y}</b><br>\u20ac%{x:,.0f} gross revenue<extra></extra>",
        ))
        l_rev = base_layout("Gross Revenue by SKU", height=max(300, len(sku_prim) * 36))
        l_rev["xaxis"]["tickprefix"] = "\u20ac"
        l_rev["margin"]["r"] = 80
        l_rev["margin"]["l"] = 220
        fig_rev.update_layout(**l_rev)
        st.plotly_chart(fig_rev, use_container_width=True)

        # ── Net contribution per SKU (revenue minus allocated expenses) ────────
        st.markdown("<h2>Revenue After Expenses \u2014 Estimated Contribution per SKU</h2>",
                    unsafe_allow_html=True)
        st.markdown(
            f"<p style='font-size:12px;color:{MID};margin-top:-6px;'>"
            "Total expenses allocated to each SKU by its share of gross revenue. "
            "Shows which SKUs generate real margin after all costs.</p>",
            unsafe_allow_html=True,
        )

        if not exp_df.empty:
            # Mirror the profitability waterfall logic exactly:
            # stock_val removed from COGS (balance sheet); promo + internal added as extra lines
            _prod_cost_raw = exp_df["production_cost"].sum()
            _total_exp_sku = (max(_prod_cost_raw - xl_stock_val, 0)
                              + xl_promo_val + xl_internal_val
                              + exp_df["logistics"].sum()
                              + exp_df["marketing_promo"].sum())
            _total_rev_sku = sku_prim["gross"].sum()
            sku_contrib = sku_prim.copy()
            sku_contrib["rev_share"]    = sku_contrib["gross"] / max(_total_rev_sku, 1)
            sku_contrib["alloc_cost"]   = (sku_contrib["rev_share"] * _total_exp_sku).round(2)
            sku_contrib["contribution"] = (sku_contrib["gross"] - sku_contrib["alloc_cost"]).round(2)
            sku_contrib["contrib_pct"]  = (
                sku_contrib["contribution"] / sku_contrib["gross"].clip(lower=0.01) * 100
            ).round(1)
            sku_contrib_sorted = sku_contrib.sort_values("contribution", ascending=True)

            _contrib_colors = [GREEN if v >= 0 else CORAL for v in sku_contrib_sorted["contribution"]]
            fig_contrib = go.Figure(go.Bar(
                x=sku_contrib_sorted["contribution"],
                y=sku_contrib_sorted["sku_name"],
                orientation="h",
                marker=dict(color=_contrib_colors),
                text=[f"\u20ac{v:,.0f}" for v in sku_contrib_sorted["contribution"]],
                textposition="outside",
                hovertemplate="<b>%{y}</b><br>Contribution: \u20ac%{x:,.0f}<extra></extra>",
            ))
            l_contrib = base_layout(
                "Estimated Contribution after Expenses (revenue \u2212 allocated costs)",
                height=max(300, len(sku_contrib) * 36),
            )
            l_contrib["xaxis"]["tickprefix"] = "\u20ac"
            l_contrib["xaxis"]["zeroline"] = True
            l_contrib["xaxis"]["zerolinecolor"] = BROWN
            l_contrib["margin"]["r"] = 80
            l_contrib["margin"]["l"] = 220
            fig_contrib.update_layout(**l_contrib)
            st.plotly_chart(fig_contrib, use_container_width=True)

            _sku_out = sku_contrib[["sku_name","bottles","gross","alloc_cost","contribution","contrib_pct"]].copy()
            _sku_out.columns = ["SKU","Pieces","Revenue (\u20ac)","Allocated Cost (\u20ac)","Contribution (\u20ac)","Margin %"]
            _sku_out = _sku_out.sort_values("Contribution (\u20ac)", ascending=False)
            _sku_out["Pieces"]                  = _sku_out["Pieces"].apply(lambda x: f"{int(x):,}")
            _sku_out["Revenue (\u20ac)"]         = _sku_out["Revenue (\u20ac)"].apply(lambda x: f"\u20ac{x:,.0f}")
            _sku_out["Allocated Cost (\u20ac)"]  = _sku_out["Allocated Cost (\u20ac)"].apply(lambda x: f"\u20ac{x:,.0f}")
            _sku_out["Contribution (\u20ac)"]    = _sku_out["Contribution (\u20ac)"].apply(lambda x: f"\u20ac{x:,.0f}")
            _sku_out["Margin %"]                = _sku_out["Margin %"].apply(lambda x: f"{x:.1f}%")
            st.dataframe(_sku_out, use_container_width=True, hide_index=True)
        else:
            _sku_simple = sku_prim[["sku_name","bottles","gross"]].copy()
            _sku_simple.columns = ["SKU","Pieces","Gross Revenue (\u20ac)"]
            _sku_simple["Pieces"]                 = _sku_simple["Pieces"].apply(lambda x: f"{int(x):,}")
            _sku_simple["Gross Revenue (\u20ac)"]  = _sku_simple["Gross Revenue (\u20ac)"].apply(lambda x: f"\u20ac{x:,.0f}")
            st.caption("Upload an Excel file with an Expenses sheet to see cost-adjusted contribution per SKU.")
            st.dataframe(_sku_simple, use_container_width=True, hide_index=True)
    else:
        st.info("No primary sales data available.")


# ──────────────────────────────────────────────────────────────────────────────
# FOOTER
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown(
    f"<p style='text-align:center; color:{MID}; font-size:10px;'>"
    "Niamito Business Intelligence · "
    "Data: Niamito_Master_Tables.xlsx · "
    "Built with Streamlit"
    "</p>",
    unsafe_allow_html=True,
)
